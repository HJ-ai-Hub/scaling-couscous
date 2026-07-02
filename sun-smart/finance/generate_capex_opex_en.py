"""
Generates SUN_SMART_CAPEX_OPEX_EN.xlsx (English version) — a full CAPEX / OPEX /
cash flow model for SUN SMART Sdn Bhd, built from the assumptions and cost
items documented in Project A OS Master Document V0.7 (chapters 1-3, 5, 6, 8,
Decision Log).

All figures are editable assumptions (yellow input cells) with formulas
recalculating totals, break-even and 12-month cash flow automatically.
Currency: MYR (RM).

Run: python3 generate_capex_opex_en.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ---------- shared styles ----------
NAVY = "1F3864"
BLUE = "2A78D6"
LIGHT_BLUE = "DCE9F9"
YELLOW = "FFF7CC"
GREY = "F2F2F0"
GOOD = "0CA30C"
CRIT = "D03B3B"
WHITE = "FFFFFF"

title_font = Font(name="Calibri", size=16, bold=True, color=WHITE)
title_fill = PatternFill("solid", fgColor=NAVY)
section_font = Font(name="Calibri", size=12, bold=True, color=WHITE)
section_fill = PatternFill("solid", fgColor=BLUE)
header_font = Font(name="Calibri", size=10, bold=True, color=NAVY)
header_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
input_fill = PatternFill("solid", fgColor=YELLOW)
total_font = Font(name="Calibri", size=10, bold=True)
total_fill = PatternFill("solid", fgColor=GREY)
note_font = Font(name="Calibri", size=9, italic=True, color="666666")
thin = Side(style="thin", color="C9C9C9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
RM = '"RM" #,##0.00'
RM0 = '"RM" #,##0'
PCT = '0.0%'


def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_section(ws, row, col, span, text):
    start = ws.cell(row=row, column=col)
    end_col = col + span - 1
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    start.value = text
    start.font = section_font
    start.fill = section_fill
    start.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_header_row(ws, row, col_start, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def input_cell(ws, coord, value=None, number_format=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.fill = input_fill
    c.border = border
    if number_format:
        c.number_format = number_format
    return c


# =====================================================================
# SHEET 0: Instructions
# =====================================================================
ws = wb.active
ws.title = "Instructions"
set_col_widths(ws, [4, 100])
style_title(ws, "A1:B1", "SUN SMART · CAPEX / OPEX Financial Model")
ws["B3"] = ("Built from the Project A OS · SUN SMART Master Document V0.7 (Chapters 1-3, 5, 6, 8 "
            "and the Decision Log). Fills the gap flagged in the document as \"Chapter 6 - Financial "
            "Budget - Not yet completed\".")
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 40

rows = [
    ("Yellow cells", "Editable assumptions/inputs. Once you have real quotes, edit these cells and every total recalculates automatically."),
    ("White/grey cells", "Calculated automatically by formula - do not need to be edited."),
    ("CAPEX - Startup Capital Expenditure", "One-off costs before opening: deposits, renovation, equipment, initial inventory, setup fees."),
    ("OPEX - Monthly Operating Expenses", "Recurring monthly costs after opening: rent, salary, utilities, payment fees, etc."),
    ("Cash Flow & P&L Forecast", "12-month projection of revenue, gross margin, expenses and net cash flow, including a break-even indicator."),
    ("Key Assumptions", "All key assumptions (margins, revenue mix, footfall conversion, etc.) in one place; every other sheet references this sheet."),
]
r = 5
for label, desc in rows:
    ws.cell(row=r, column=2, value=f"* {label}: {desc}").alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 1

ws.cell(row=r + 1, column=2, value="How to use this workbook:").font = Font(bold=True)
steps = [
    "1. Start by entering any real numbers you already have (rent quotes, supplier prices, etc.) into the \"Key Assumptions\" sheet. Where you don't have real numbers yet, keep the estimates from the master document.",
    "2. \"CAPEX - Startup Capital Expenditure\" and \"OPEX - Monthly Operating Expenses\" automatically reference the Key Assumptions sheet, but you can also override individual line items directly on those two sheets.",
    "3. \"12-Month Cash Flow Forecast\" rolls everything up automatically, so you can check whether the RM50,000-100,000 startup budget is enough to survive the \"3 consecutive months of profit + positive cash flow\" validation bar set in Decision Log #024.",
    "4. Whenever you get a new quote (e.g. mall Level B rental, supplier Level B pricing), come back and update the corresponding cell so this stays the single master financial file, in line with Decision Log #026 (Section 10.2) of the master document.",
]
for i, s in enumerate(steps):
    ws.cell(row=r + 2 + i, column=2, value=s).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r + 2 + i].height = 30

ws.sheet_view.showGridLines = False

# =====================================================================
# SHEET 1: Key Assumptions — single source of truth
# AC = name -> cell coordinate dict, built as we go (no hand-guessed rows)
# =====================================================================
ws = wb.create_sheet("Key Assumptions")
set_col_widths(ws, [3, 38, 16, 46])
style_title(ws, "A1:D1", "Key Assumptions (Yellow = editable; every other sheet references this sheet)")

AC = {}  # name -> "C{row}"
ASSUMP_SHEET = "Key Assumptions"


def add_assumption_block(start_row, title, items):
    style_section(ws, start_row, 2, 3, title)
    r = start_row + 1
    for name, val, fmt, note in items:
        ws.cell(row=r, column=2, value=name).border = border
        input_cell(ws, f"C{r}", val, fmt)
        c = ws.cell(row=r, column=4, value=note)
        c.font = note_font
        c.alignment = Alignment(wrap_text=True)
        AC[name] = f"C{r}"
        r += 1
    return r + 1


r = 3
r = add_assumption_block(r, "Startup Budget & Launch", [
    ("Startup budget - lower bound", 50000, RM0, "Doc §3.5 / Decision Log #018: RM50,000-100,000"),
    ("Startup budget - upper bound", 100000, RM0, ""),
])

r = add_assumption_block(r, "Revenue Mix Assumptions (Doc §2.3 / Decision Log #005)", [
    ("New phone revenue share", 0.45, PCT, ""),
    ("Used phone revenue share", 0.45, PCT, ""),
    ("Accessories revenue share", 0.10, PCT, ""),
])

r = add_assumption_block(r, "Gross Margin Assumptions (Doc §2.4 / Decision Log #006)", [
    ("New phone gross margin", 0.025, PCT, "Doc range 2%-3%, midpoint used"),
    ("Used phone gross margin", 0.25, PCT, ""),
    ("Accessories gross margin", 0.40, PCT, "Doc range 30%-50%, midpoint used, pending supplier quotes"),
])

r = add_assumption_block(r, "Inventory Capital Allocation (Doc §3.6 / Decision Log #008)", [
    ("New phone inventory share", 0.70, PCT, ""),
    ("Used phone inventory share", 0.20, PCT, ""),
    ("Accessories inventory share", 0.10, PCT, ""),
    ("Initial inventory budget total", 35000, RM0, "Recommended to not exceed 50-60% of startup budget; buy in batches, see Doc §3.6.4"),
])

r = add_assumption_block(r, "Sales Volume Assumptions (Month 1 baseline; Month 2+ grows at monthly growth rate)", [
    ("Avg monthly phone units sold", 40, "0", "Conservative 0-to-1 stage assumption; calibrate with real Month 1 data"),
    ("Avg selling price per phone", 1400, RM0, "Near the midpoint of the doc's RM500-3000 product range"),
    ("Avg monthly accessory orders", 150, "0", ""),
    ("Avg accessory order value", 45, RM0, ""),
    ("Month-over-month sales growth rate", 0.03, PCT, "Assumes slow growth driven by word-of-mouth and online traffic"),
])

r = add_assumption_block(r, "Payment Channel Mix & Fee Assumptions (Doc §5.6)", [
    ("Cash / bank transfer share", 0.45, PCT, "No fee"),
    ("DuitNow QR share", 0.20, PCT, ""),
    ("Credit card share", 0.15, PCT, ""),
    ("PayLater share", 0.20, PCT, "Doc range 3-6%, upper-middle used"),
    ("DuitNow QR fee rate", 0.010, PCT, ""),
    ("Credit card fee rate", 0.015, PCT, ""),
    ("PayLater fee rate", 0.045, PCT, ""),
])

r = add_assumption_block(r, "Staffing Assumptions (Doc §2.14)", [
    ("Number of staff", 1, "0", "1 sales staff at launch"),
    ("Staff base salary", 2000, RM0, ""),
    ("Employer EPF SOCSO EIS rate", 0.13, PCT, "Approx. 13%"),
    ("Sales commission pct of gross margin", 0.03, PCT, "Assumed value, to be agreed with staff separately"),
])

ws.sheet_view.showGridLines = False


def A(name):
    """Reference an assumption cell by its label, e.g. A('New phone gross margin') -> "'Key Assumptions'!C13" """
    return f"'{ASSUMP_SHEET}'!{AC[name]}"


# =====================================================================
# SHEET 2: CAPEX - Startup Capital Expenditure
# =====================================================================
ws = wb.create_sheet("CAPEX Startup Capex")
set_col_widths(ws, [3, 38, 14, 14, 14, 46])
style_title(ws, "A1:F1", "CAPEX · Startup Capital Expenditure (maps to Doc Chapter 3 Checklist)")

subtotal_cells = []
cash_reserve_capex_row = None


def capex_section(start_row, title, items):
    """items: (name, qty, unit_cost, note); returns (next_row, subtotal_coord)"""
    style_section(ws, start_row, 2, 5, title)
    r = start_row + 1
    style_header_row(ws, r, 2, ["Item", "Qty", "Unit Cost (RM)", "Subtotal (RM)", "Note"])
    r += 1
    first_data_row = r
    for name, qty, unit_cost, note in items:
        ws.cell(row=r, column=2, value=name).border = border
        input_cell(ws, f"C{r}", qty, "0")
        input_cell(ws, f"D{r}", unit_cost, RM0)
        f = ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
        f.number_format = RM
        f.border = border
        n = ws.cell(row=r, column=6, value=note)
        n.font = note_font
        n.alignment = Alignment(wrap_text=True)
        r += 1
    last_data_row = r - 1
    ws.cell(row=r, column=2, value="Subtotal").font = total_font
    ws.cell(row=r, column=2).fill = total_fill
    for col in (3, 4):
        ws.cell(row=r, column=col).fill = total_fill
    tot = ws.cell(row=r, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})")
    tot.number_format = RM
    tot.font = total_font
    tot.fill = total_fill
    return r + 2, tot.coordinate


def capex_formula_section(start_row, title, rows):
    """rows: (name, formula, note); returns (next_row, subtotal_coord)"""
    style_section(ws, start_row, 2, 5, title)
    rr = start_row + 1
    style_header_row(ws, rr, 2, ["Item", "Qty", "Unit Cost (RM)", "Subtotal (RM)", "Note"])
    rr += 1
    first = rr
    for name, formula, note in rows:
        ws.cell(row=rr, column=2, value=name).border = border
        ws.cell(row=rr, column=3).border = border
        ws.cell(row=rr, column=4).border = border
        c = ws.cell(row=rr, column=5, value=formula)
        c.number_format = RM
        c.border = border
        n = ws.cell(row=rr, column=6, value=note)
        n.font = note_font
        n.alignment = Alignment(wrap_text=True)
        rr += 1
    last = rr - 1
    ws.cell(row=rr, column=2, value="Subtotal").font = total_font
    ws.cell(row=rr, column=2).fill = total_fill
    ws.cell(row=rr, column=3).fill = total_fill
    ws.cell(row=rr, column=4).fill = total_fill
    tot = ws.cell(row=rr, column=5, value=f"=SUM(E{first}:E{last})")
    tot.number_format = RM
    tot.font = total_font
    tot.fill = total_fill
    return rr + 2, tot.coordinate


r = 3
r, t = capex_section(r, "1. Mall Lease - One-off Costs (Doc §3.4)", [
    ("Rental deposit (typically 2-3 months' rent)", 1, 6000, "Pending mall Level B quote, see Doc §7.6 questions checklist"),
    ("Utilities deposit", 1, 500, ""),
    ("Renovation deposit", 1, 1000, ""),
    ("Signboard application / drawing approval fee", 1, 800, ""),
    ("Legal fee / stamp duty (tenancy agreement)", 1, 500, ""),
])
subtotal_cells.append(("Mall lease one-off costs", t))

r, t = capex_section(r, "2. Renovation (Doc §3.8)", [
    ("Display cabinets (movable / reusable)", 3, 800, "Doc emphasises movable, reusable fixtures for future relocation"),
    ("Counter / cashier station", 1, 1200, ""),
    ("Signboard production & installation", 1, 2500, ""),
    ("Lighting works", 1, 1000, ""),
    ("Power / network wiring", 1, 800, ""),
    ("Basic paint and flooring touch-up", 1, 1500, ""),
    ("Renovation - other misc", 1, 1000, ""),
])
subtotal_cells.append(("Renovation", t))

r, t = capex_section(r, "3. Shop Equipment (Doc §3.9)", [
    ("CCTV system", 1, 900, ""),
    ("Computer / laptop", 1, 2200, ""),
    ("Printer", 1, 400, ""),
    ("Cash drawer", 1, 250, ""),
    ("Wi-Fi router", 1, 200, ""),
    ("POS / barcode scanner (if adopted)", 1, 600, "Doc §4.4: Phase 1 uses Google Sheets, POS deferred - can be 0"),
    ("Anti-theft display stand / small safe", 1, 500, ""),
    ("Fire extinguisher / UPS / other safety equipment", 1, 300, ""),
    ("Packaging bags / receipt books / label printer, etc.", 1, 500, ""),
])
subtotal_cells.append(("Shop equipment", t))

r, t = capex_section(r, "4. Payments & Systems Setup (Doc §3.10, 5.6)", [
    ("DuitNow QR application / activation fee", 1, 0, "Usually free, pending Level B confirmation"),
    ("Credit card terminal setup fee", 1, 300, "Pending HitPay / Revenue Monster quotes, see Doc §5.8"),
    ("PayLater platform setup fee", 1, 300, ""),
    ("WhatsApp Business setup (labour time, can be 0 if no cash cost)", 1, 0, ""),
    ("Website build cost", 1, 3000, "Doc §3.11.2: if online payment can't be ready in time, use WhatsApp ordering first"),
])
subtotal_cells.append(("Payments & systems setup", t))

r, t = capex_section(r, "5. Company & Compliance Setup (mirrors the not-yet-completed Chapter 9 - Legal & Documentation)", [
    ("SSM annual fee / filing (if renewal due)", 1, 300, "Doc §3.2: SSM registration already done; this is an ongoing annual maintenance reminder"),
    ("Business Premise License", 1, 500, "Local council licence, fee depends on shop location"),
    ("Signboard License", 1, 300, ""),
    ("Initial legal / contract review (tenancy, partnership agreement)", 1, 1500, "Maps to Doc §2.15 written partner agreement and the Chapter 9 legal gap"),
    ("First-year insurance premium (if purchased)", 1, 0, "Doc §3.14: not purchasing immediately, research first; fill in once decided"),
])
subtotal_cells.append(("Company & compliance setup", t))

r, t = capex_formula_section(r, "6. Initial Inventory (Doc §3.6 / Decision Log #008, auto-calculated on the 70/20/10 split)", [
    ("New phone initial inventory", f"={A('Initial inventory budget total')}*{A('New phone inventory share')}", "Formula: initial inventory budget x new phone share (Key Assumptions sheet)"),
    ("Used phone initial inventory", f"={A('Initial inventory budget total')}*{A('Used phone inventory share')}", "Formula: initial inventory budget x used phone share"),
    ("Accessories initial inventory", f"={A('Initial inventory budget total')}*{A('Accessories inventory share')}", "Formula: initial inventory budget x accessories share"),
])
subtotal_cells.append(("Initial inventory", t))

r, t = capex_section(r, "7. Cash Flow Reserve (Doc §3.5 cash-flow-first principle / Decision Log #018)", [
    ("Emergency / operating cash reserve (auto-linked to the OPEX sheet's recommendation, see note)", 1, 0, "Do not spend the entire budget by opening day; this line becomes a formula once the OPEX sheet exists"),
])
cash_reserve_capex_row = r - 2  # row of the single item above
subtotal_cells.append(("Cash flow reserve", t))

gt_row = r + 1
style_section(ws, gt_row, 2, 5, "CAPEX Grand Total")
gt_row += 1
ws.cell(row=gt_row, column=2, value="CAPEX Grand Total").font = Font(bold=True, size=12)
formula_parts = "+".join(coord for _, coord in subtotal_cells)
grand = ws.cell(row=gt_row, column=5, value=f"={formula_parts}")
grand.number_format = RM
grand.font = Font(bold=True, size=12)
grand.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws.cell(row=gt_row, column=2).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

gt_row += 1
ws.cell(row=gt_row, column=2, value="vs. startup budget lower bound").font = Font(italic=True)
diff = ws.cell(row=gt_row, column=5, value=f"={A('Startup budget - lower bound')}-{grand.coordinate}")
diff.number_format = RM
diff.font = Font(italic=True)

gt_row += 1
ws.cell(row=gt_row, column=2, value="vs. startup budget upper bound").font = Font(italic=True)
diff2 = ws.cell(row=gt_row, column=5, value=f"={A('Startup budget - upper bound')}-{grand.coordinate}")
diff2.number_format = RM
diff2.font = Font(italic=True)

CAPEX_TOTAL_CELL = grand.coordinate
CAPEX_SHEET = "CAPEX Startup Capex"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B3"

# =====================================================================
# SHEET 3: OPEX - Monthly Operating Expenses
# =====================================================================
ws = wb.create_sheet("OPEX Monthly Opex")
set_col_widths(ws, [3, 38, 14, 46])
style_title(ws, "A1:D1", "OPEX · Monthly Operating Expenses (maps to Doc §2.16 core cost structure)")

opex_subtotals = []


def opex_section(start_row, title, items):
    """items: (name, value_or_formula, note, is_formula)"""
    style_section(ws, start_row, 2, 3, title)
    r = start_row + 1
    style_header_row(ws, r, 2, ["Item", "Monthly Amount (RM)", "Note"])
    r += 1
    first = r
    for name, val, note, is_formula in items:
        ws.cell(row=r, column=2, value=name).border = border
        c = ws.cell(row=r, column=3, value=val)
        c.number_format = RM
        c.border = border
        if not is_formula:
            c.fill = input_fill
        n = ws.cell(row=r, column=4, value=note)
        n.font = note_font
        n.alignment = Alignment(wrap_text=True)
        r += 1
    last = r - 1
    ws.cell(row=r, column=2, value="Subtotal").font = total_font
    ws.cell(row=r, column=2).fill = total_fill
    ws.cell(row=r, column=3).fill = total_fill
    tot = ws.cell(row=r, column=3, value=f"=SUM(C{first}:C{last})")
    tot.number_format = RM
    tot.font = total_font
    tot.fill = total_fill
    return r + 2, tot.coordinate


r = 3
r, t = opex_section(r, "1. Premises (Doc §3.4)", [
    ("Base mall rent", 3500, "Pending mall Level B quote, see Doc §7.9 contact list", False),
    ("Management / promotion fee", 400, "", False),
    ("Utilities (electricity/water)", 500, "", False),
    ("Air-conditioning charge (if billed separately)", 200, "", False),
    ("Internet / telecom", 150, "", False),
    ("Refuse collection / other misc fees", 100, "", False),
])
opex_subtotals.append(("Premises", t))

r, t = opex_section(r, "2. Staffing Costs (Doc §2.16 / auto-linked to Key Assumptions)", [
    ("Staff base salary", f"={A('Staff base salary')}", "Linked to Key Assumptions - staff base salary", True),
    ("EPF/SOCSO/EIS employer portion", f"={A('Staff base salary')}*{A('Employer EPF SOCSO EIS rate')}", "Formula: salary x employer contribution rate", True),
    ("Sales commission (estimate; actual linked to gross margin in forecast sheet)", 300, "Initial estimate, calibrate with first month's actual gross margin", False),
])
opex_subtotals.append(("Staffing costs", t))

r, t = opex_section(r, "3. Payment Channel Fees (variable cost, auto-calculated from revenue)", [
    ("Note: fees are a variable cost, already auto-calculated by channel mix and fee rate in \"12-Month Cash Flow Forecast\"; kept at 0 here to avoid double-counting", 0, "See the \"12-Month Cash Flow Forecast\" sheet for the monthly fee amount", False),
])
opex_subtotals.append(("Payment channel fees", t))

r, t = opex_section(r, "4. Marketing & Customer Acquisition (Doc §2.13, 3.15)", [
    ("Facebook ads", 500, "", False),
    ("TikTok content / ads", 300, "", False),
    ("Launch / promo materials (reverts to a small standing budget after month 3)", 200, "", False),
])
opex_subtotals.append(("Marketing & customer acquisition", t))

r, t = opex_section(r, "5. Systems & Subscriptions (Doc §4.3-4.4, 3.11)", [
    ("Google Sheets (official inventory/records system, free-~RM300/yr pro-rated monthly)", 25, "Doc §4.4: Decision Log #007 confirms the free/low-cost route", False),
    ("Website maintenance / domain / hosting", 100, "", False),
    ("WhatsApp Business (free; add cost if using a 3rd-party tool)", 0, "", False),
])
opex_subtotals.append(("Systems & subscriptions", t))

r, t = opex_section(r, "6. After-sales & Risk Reserve (Doc §2.12, 3.13, 3.14 / mirrors the Chapter 9 legal gap)", [
    ("Repair / after-sales handling reserve", 300, "For external repair-partner fees and non-warranty repair handling", False),
    ("Insurance allocation (if purchased, see Doc §3.14)", 0, "Not purchasing yet, researching first; fill in annual premium / 12 once decided", False),
    ("Bad debt / return-loss reserve (approx. 0.5% of revenue, estimate)", 150, "", False),
])
opex_subtotals.append(("After-sales & risk reserve", t))

r, t = opex_section(r, "7. Other Admin (mirrors the not-yet-completed Chapter 9 - Legal & Documentation)", [
    ("Accounting / bookkeeping service", 250, "", False),
    ("Bank charges / other admin misc", 50, "", False),
    ("LHDN e-Invoice / SST compliance reserve (can be 0 before threshold is reached)", 0, "Reserve once revenue nears the SST threshold (RM500k) or e-Invoice mandatory phase", False),
])
opex_subtotals.append(("Other admin", t))

gt_row = r + 1
style_section(ws, gt_row, 2, 3, "OPEX Grand Total")
gt_row += 1
ws.cell(row=gt_row, column=2, value="OPEX Grand Total").font = Font(bold=True, size=12)
formula_parts = "+".join(coord for _, coord in opex_subtotals)
opex_grand = ws.cell(row=gt_row, column=3, value=f"={formula_parts}")
opex_grand.number_format = RM
opex_grand.font = Font(bold=True, size=12)
opex_grand.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws.cell(row=gt_row, column=2).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

gt_row += 1
ws.cell(row=gt_row, column=2, value="Recommended cash reserve (2.5 months of fixed costs; Doc §3.5 requires 2-3 months)").font = Font(italic=True)
reserve = ws.cell(row=gt_row, column=3, value=f"={opex_grand.coordinate}*2.5")
reserve.number_format = RM
reserve.font = Font(italic=True)

OPEX_TOTAL_CELL = opex_grand.coordinate
OPEX_SHEET = "OPEX Monthly Opex"
CASH_RESERVE_CELL = reserve.coordinate
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B3"

# now wire CAPEX's cash-reserve line to this OPEX suggestion
ws_capex = wb[CAPEX_SHEET]
e_cell = ws_capex.cell(row=cash_reserve_capex_row, column=5)
e_cell.value = f"='{OPEX_SHEET}'!{CASH_RESERVE_CELL}"
e_cell.number_format = RM

# =====================================================================
# SHEET 4: 12-Month Cash Flow Forecast
# =====================================================================
ws = wb.create_sheet("12-Month Cash Flow Forecast")
set_col_widths(ws, [34] + [12] * 12 + [14])
style_title(ws, "A1:N1", "12-Month Cash Flow & P&L Forecast (auto-linked to Key Assumptions and OPEX; Month 2+ compounds at the growth rate)")

months = [f"Month {i}" for i in range(1, 13)]
style_header_row(ws, 3, 1, ["Line Item"] + months + ["Total"])


def col_letter_for_month(i):  # i=0..11 -> columns B..M
    return get_column_letter(2 + i)


row = 4
ws.cell(row=row, column=1, value="Phone units sold (compounds from Month 2 at growth rate)")
for i in range(12):
    if i == 0:
        c = ws.cell(row=row, column=2, value=f"={A('Avg monthly phone units sold')}")
    else:
        prev = col_letter_for_month(i - 1)
        c = ws.cell(row=row, column=2 + i, value=f"={prev}{row}*(1+{A('Month-over-month sales growth rate')})")
    c.number_format = "0.0"
qty_row = row

row += 1
ws.cell(row=row, column=1, value="Avg selling price per phone (RM)")
for i in range(12):
    ws.cell(row=row, column=2 + i, value=f"={A('Avg selling price per phone')}").number_format = RM
price_row = row

row += 1
ws.cell(row=row, column=1, value="Phone sales revenue (RM)")
for i in range(12):
    col = col_letter_for_month(i)
    ws.cell(row=row, column=2 + i, value=f"={col}{qty_row}*{col}{price_row}").number_format = RM
phone_rev_row = row

row += 1
ws.cell(row=row, column=1, value="Accessory orders")
for i in range(12):
    if i == 0:
        c = ws.cell(row=row, column=2, value=f"={A('Avg monthly accessory orders')}")
    else:
        prev = col_letter_for_month(i - 1)
        c = ws.cell(row=row, column=2 + i, value=f"={prev}{row}*(1+{A('Month-over-month sales growth rate')})")
    c.number_format = "0.0"
acc_qty_row = row

row += 1
ws.cell(row=row, column=1, value="Avg accessory order value (RM)")
for i in range(12):
    ws.cell(row=row, column=2 + i, value=f"={A('Avg accessory order value')}").number_format = RM
acc_price_row = row

row += 1
ws.cell(row=row, column=1, value="Accessories sales revenue (RM)")
for i in range(12):
    col = col_letter_for_month(i)
    ws.cell(row=row, column=2 + i, value=f"={col}{acc_qty_row}*{col}{acc_price_row}").number_format = RM
acc_rev_row = row

row += 1
ws.cell(row=row, column=1, value="Total revenue (RM)").font = total_font
for i in range(12):
    col = col_letter_for_month(i)
    c = ws.cell(row=row, column=2 + i, value=f"={col}{phone_rev_row}+{col}{acc_rev_row}")
    c.number_format = RM
    c.font = total_font
    c.fill = total_fill
total_rev_row = row

row += 1
ws.cell(row=row, column=1, value="Estimated gross margin (RM) - new/used/accessories blended")
new_share = A('New phone revenue share')
used_share = A('Used phone revenue share')
new_margin = A('New phone gross margin')
used_margin = A('Used phone gross margin')
acc_margin = A('Accessories gross margin')
for i in range(12):
    col = col_letter_for_month(i)
    phone_margin_formula = f"({col}{phone_rev_row}*(({new_share}/({new_share}+{used_share}))*{new_margin}+({used_share}/({new_share}+{used_share}))*{used_margin}))"
    acc_margin_formula = f"({col}{acc_rev_row}*{acc_margin})"
    c = ws.cell(row=row, column=2 + i, value=f"={phone_margin_formula}+{acc_margin_formula}")
    c.number_format = RM
gross_margin_row = row

row += 1
ws.cell(row=row, column=1, value="Payment processing fees (RM)")
cash_share = A('Cash / bank transfer share')
duitnow_share = A('DuitNow QR share')
card_share = A('Credit card share')
paylater_share = A('PayLater share')
duitnow_fee = A('DuitNow QR fee rate')
card_fee = A('Credit card fee rate')
paylater_fee = A('PayLater fee rate')
for i in range(12):
    col = col_letter_for_month(i)
    fee = f"={col}{total_rev_row}*({duitnow_share}*{duitnow_fee}+{card_share}*{card_fee}+{paylater_share}*{paylater_fee})"
    ws.cell(row=row, column=2 + i, value=fee).number_format = RM
fee_row = row

row += 1
ws.cell(row=row, column=1, value="Monthly fixed OPEX (RM)")
for i in range(12):
    ws.cell(row=row, column=2 + i, value=f"='{OPEX_SHEET}'!{OPEX_TOTAL_CELL}").number_format = RM
opex_row = row

row += 1
ws.cell(row=row, column=1, value="Sales commission (% of gross margin)")
for i in range(12):
    col = col_letter_for_month(i)
    ws.cell(row=row, column=2 + i, value=f"={col}{gross_margin_row}*{A('Sales commission pct of gross margin')}").number_format = RM
commission_row = row

row += 1
ws.cell(row=row, column=1, value="Net profit (RM) = Gross margin - fees - OPEX - commission").font = total_font
for i in range(12):
    col = col_letter_for_month(i)
    c = ws.cell(row=row, column=2 + i, value=f"={col}{gross_margin_row}-{col}{fee_row}-{col}{opex_row}-{col}{commission_row}")
    c.number_format = RM
    c.font = total_font
    c.fill = total_fill
net_profit_row = row

row += 1
ws.cell(row=row, column=1, value="Cumulative net cash flow (RM, starts at 0, excludes CAPEX)")
for i in range(12):
    col = col_letter_for_month(i)
    if i == 0:
        ws.cell(row=row, column=2 + i, value=f"={col}{net_profit_row}").number_format = RM
    else:
        prev = col_letter_for_month(i - 1)
        ws.cell(row=row, column=2 + i, value=f"={prev}{row}+{col}{net_profit_row}").number_format = RM
cum_row = row

# totals column (N)
n_col = 14
for r_ in [qty_row, phone_rev_row, acc_qty_row, acc_rev_row, total_rev_row, gross_margin_row, fee_row, opex_row, commission_row, net_profit_row]:
    c = ws.cell(row=r_, column=n_col, value=f"=SUM(B{r_}:M{r_})")
    c.number_format = "0.0" if r_ in (qty_row, acc_qty_row) else RM
    c.font = total_font

ws.cell(row=3, column=n_col, value="12-Month Total").font = header_font
ws.cell(row=3, column=n_col).fill = header_fill
ws.cell(row=3, column=n_col).border = border

row += 2
ws.cell(row=row, column=1, value="Decision Log #024 validation bar: 3 consecutive months of net profit > 0 and positive cumulative net cash flow").font = Font(italic=True, size=9, color="666666")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)

red_font = Font(color=CRIT)
green_font = Font(color=GOOD)
ws.conditional_formatting.add(f"B{net_profit_row}:M{net_profit_row}", CellIsRule(operator='lessThan', formula=['0'], font=red_font))
ws.conditional_formatting.add(f"B{net_profit_row}:M{net_profit_row}", CellIsRule(operator='greaterThanOrEqual', formula=['0'], font=green_font))
ws.conditional_formatting.add(f"B{cum_row}:N{cum_row}", CellIsRule(operator='lessThan', formula=['0'], font=red_font))
ws.conditional_formatting.add(f"B{cum_row}:N{cum_row}", CellIsRule(operator='greaterThanOrEqual', formula=['0'], font=green_font))

ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"
FORECAST_SHEET = "12-Month Cash Flow Forecast"
FORECAST_TOTAL_REV_CELL = f"N{total_rev_row}"
FORECAST_TOTAL_NETPROFIT_CELL = f"N{net_profit_row}"
FORECAST_LAST_CUM_CELL = f"M{cum_row}"

# =====================================================================
# SHEET 5: Summary Dashboard
# =====================================================================
ws = wb.create_sheet("Summary")
wb.move_sheet("Summary", offset=-5)  # place right after Instructions, before Key Assumptions
set_col_widths(ws, [4, 38, 20, 46])
style_title(ws, "A1:D1", "SUN SMART · CAPEX/OPEX Summary")

ws["B3"] = "Currency: Malaysian Ringgit (RM)   |   Single source of truth for all figures: \"Key Assumptions\" sheet"
ws["B3"].font = note_font

r = 5
style_section(ws, r, 2, 2, "One-off Capital Expenditure (CAPEX)")
r += 1
ws.cell(row=r, column=2, value="CAPEX total (incl. cash reserve)")
c = ws.cell(row=r, column=3, value=f"='{CAPEX_SHEET}'!{CAPEX_TOTAL_CELL}")
c.number_format = RM
c.font = Font(bold=True, size=13)
r += 2

style_section(ws, r, 2, 2, "Monthly Operating Expenses (OPEX)")
r += 1
ws.cell(row=r, column=2, value="Monthly OPEX total")
c = ws.cell(row=r, column=3, value=f"='{OPEX_SHEET}'!{OPEX_TOTAL_CELL}")
c.number_format = RM
c.font = Font(bold=True, size=13)
r += 1
ws.cell(row=r, column=2, value="Recommended cash reserve (2.5 months)")
c = ws.cell(row=r, column=3, value=f"='{OPEX_SHEET}'!{CASH_RESERVE_CELL}")
c.number_format = RM
r += 2

style_section(ws, r, 2, 2, "Total Startup Funding Requirement")
r += 1
ws.cell(row=r, column=2, value="CAPEX total (already incl. cash reserve)").font = Font(bold=True)
c = ws.cell(row=r, column=3, value=f"='{CAPEX_SHEET}'!{CAPEX_TOTAL_CELL}")
c.number_format = RM
c.font = Font(bold=True, size=13)
total_need_cell = c.coordinate
r += 1
ws.cell(row=r, column=2, value="vs. RM50,000-100,000 budget range")
c = ws.cell(
    row=r, column=3,
    value=f"=IF({total_need_cell}<={A('Startup budget - lower bound')},\"Below the lower bound, ample budget\",IF({total_need_cell}<={A('Startup budget - upper bound')},\"Within the budget range\",\"Exceeds the upper bound, adjust assumptions or raise more capital\"))"
)
r += 3

style_section(ws, r, 2, 2, "12-Month Forecast Summary")
r += 1
ws.cell(row=r, column=2, value="12-month total revenue")
c = ws.cell(row=r, column=3, value=f"='{FORECAST_SHEET}'!{FORECAST_TOTAL_REV_CELL}")
c.number_format = RM
r += 1
ws.cell(row=r, column=2, value="12-month total net profit")
c = ws.cell(row=r, column=3, value=f"='{FORECAST_SHEET}'!{FORECAST_TOTAL_NETPROFIT_CELL}")
c.number_format = RM
r += 1
ws.cell(row=r, column=2, value="Cumulative net cash flow at end of Month 12")
c = ws.cell(row=r, column=3, value=f"='{FORECAST_SHEET}'!{FORECAST_LAST_CUM_CELL}")
c.number_format = RM
r += 2

note = ws.cell(row=r, column=2, value="Tip: every yellow input cell lives on the \"Key Assumptions\", \"CAPEX\" and \"OPEX\" sheets. After editing, come back to this Summary page to check you're still within budget.")
note.font = note_font
note.alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
ws.row_dimensions[r].height = 30

ws.sheet_view.showGridLines = False

# Open on the "Summary" tab by default so the CAPEX/OPEX totals are the first
# thing visible, instead of the text-only "Instructions" tab.
wb.active = wb.sheetnames.index("Summary")

OUTPUT_PATH = "SUN_SMART_CAPEX_OPEX_EN.xlsx"
wb.save(OUTPUT_PATH)

from bake_cache import bake
baked = bake(OUTPUT_PATH)
print(f"Saved {OUTPUT_PATH} (baked {baked} cached formula value(s) for non-recalculating viewers)")
print("Assumption cell map:")
for k, v in AC.items():
    print(f"  {k}: {v}")

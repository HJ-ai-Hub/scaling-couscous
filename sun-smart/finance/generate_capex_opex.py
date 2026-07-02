"""
Generates SUN_SMART_CAPEX_OPEX.xlsx — a full CAPEX / OPEX / cash flow model
for SUN SMART Sdn Bhd, built from the assumptions and cost items documented
in Project A OS Master Document V0.7 (chapters 1-3, 5, 6, 8, Decision Log).

All figures are editable assumptions (yellow input cells) with formulas
recalculating totals, break-even and 12-month cash flow automatically.
Currency: MYR (RM).

Run: python3 generate_capex_opex.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ---------- shared styles ----------
NAVY = "1F3864"
BLUE = "2A78D6"
LIGHT_BLUE = "DCE9F9"
YELLOW = "FFF7CC"
GREY = "F2F2F0"
GOOD = "0CA30C"
CRIT = "D03B3B"
WHITE = "FFFFFF"

title_font = Font(name="Calibri", size=16, bold=True, color=WHITE)
title_fill = PatternFill("solid", fgColor=NAVY)
section_font = Font(name="Calibri", size=12, bold=True, color=WHITE)
section_fill = PatternFill("solid", fgColor=BLUE)
header_font = Font(name="Calibri", size=10, bold=True, color=NAVY)
header_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
input_fill = PatternFill("solid", fgColor=YELLOW)
total_font = Font(name="Calibri", size=10, bold=True)
total_fill = PatternFill("solid", fgColor=GREY)
note_font = Font(name="Calibri", size=9, italic=True, color="666666")
thin = Side(style="thin", color="C9C9C9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
RM = '"RM" #,##0.00'
RM0 = '"RM" #,##0'
PCT = '0.0%'


def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_section(ws, row, col, span, text):
    start = ws.cell(row=row, column=col)
    end_col = col + span - 1
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    start.value = text
    start.font = section_font
    start.fill = section_fill
    start.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_header_row(ws, row, col_start, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def input_cell(ws, coord, value=None, number_format=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.fill = input_fill
    c.border = border
    if number_format:
        c.number_format = number_format
    return c


# =====================================================================
# SHEET 0: 使用说明 README
# =====================================================================
ws = wb.active
ws.title = "使用说明"
set_col_widths(ws, [4, 100])
style_title(ws, "A1:B1", "SUN SMART · CAPEX / OPEX 财务模型")
ws["B3"] = "本表根据《Project A OS · SUN SMART 经营操作系统 Master Document V0.7》第1-3、5、6、8章及 Decision Log 整理而成，用于补上文件中标注「第6章财务预算 待完成」的缺口。"
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 40

rows = [
    ("黄色底色的格子", "= 可编辑的假设/输入值。拿到真实报价后，直接改这些格子，其余公式会自动重算。"),
    ("白色/灰色格子", "= 公式自动计算，不需要手动改。"),
    ("CAPEX 一次性资本支出", "= 开店前一次性投入的钱：押金、装修、设备、首批库存、系统开办费等。"),
    ("OPEX 每月营运支出", "= 开店后每个月固定要付的钱：租金、工资、水电、支付手续费等。"),
    ("现金流与损益预测", "= 未来12个月的营收、毛利、费用、净现金流预测，含盈亏平衡点估算。"),
    ("关键假设", "= 集中列出所有关键假设（毛利率、坪效、人流转化率等），方便日后调整重跑，其余表格全部引用这张表。"),
]
r = 5
for label, desc in rows:
    ws.cell(row=r, column=2, value=f"● {label}：{desc}").alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 1

ws.cell(row=r+1, column=2, value="使用步骤：").font = Font(bold=True)
steps = [
    "1. 先在「关键假设」表填入你已经拿到的真实数字（租金、供应商报价等），没拿到的先保留文件里的估算值。",
    "2. 「CAPEX 开店资本支出」和「OPEX 每月营运支出」会自动引用「关键假设」的数字，也可以直接在两张表里覆盖个别项目。",
    "3. 「12个月现金流预测」会自动汇总，用来判断 RM50,000-100,000 的启动资金是否足够撑过 Decision Log #024 要求的『连续三个月盈利+正现金流』验证期。",
    "4. 每次拿到新报价（例如商场 Level B 租金、供应商 Level B 价格）就回来更新对应格子，保持这份表是唯一 Master 财务文件，呼应文件 10.2 的 Decision Log #026 原则。",
]
for i, s in enumerate(steps):
    ws.cell(row=r+2+i, column=2, value=s).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r+2+i].height = 30

ws.sheet_view.showGridLines = False

# =====================================================================
# SHEET 1: 关键假设 (Assumptions) — single source of truth
# AC = name -> cell coordinate dict, built as we go (no hand-guessed rows)
# =====================================================================
ws = wb.create_sheet("关键假设")
set_col_widths(ws, [3, 34, 16, 40])
style_title(ws, "A1:D1", "关键假设 Key Assumptions（黄色=可编辑，其余表格全部引用这里）")

AC = {}  # name -> "C{row}"
ASSUMP_SHEET = "关键假设"


def add_assumption_block(start_row, title, items):
    style_section(ws, start_row, 2, 3, title)
    r = start_row + 1
    for name, val, fmt, note in items:
        ws.cell(row=r, column=2, value=name).border = border
        input_cell(ws, f"C{r}", val, fmt)
        c = ws.cell(row=r, column=4, value=note)
        c.font = note_font
        c.alignment = Alignment(wrap_text=True)
        AC[name] = f"C{r}"
        r += 1
    return r + 1


r = 3
r = add_assumption_block(r, "启动预算与开业", [
    ("启动预算下限", 50000, RM0, "文件 3.5 / Decision Log #018：RM50,000-100,000"),
    ("启动预算上限", 100000, RM0, ""),
])

r = add_assumption_block(r, "收入结构假设 (文件 2.3 / Decision Log #005)", [
    ("新机收入占比", 0.45, PCT, ""),
    ("二手机收入占比", 0.45, PCT, ""),
    ("配件收入占比", 0.10, PCT, ""),
])

r = add_assumption_block(r, "毛利率假设 (文件 2.4 / Decision Log #006)", [
    ("新机毛利率", 0.025, PCT, "文件区间 2%-3%，取中间值"),
    ("二手机毛利率", 0.25, PCT, ""),
    ("配件毛利率", 0.40, PCT, "文件区间 30%-50%，取中间值，待供应商报价验证"),
])

r = add_assumption_block(r, "库存资金分配 (文件 3.6 / Decision Log #008)", [
    ("新机库存占比", 0.70, PCT, ""),
    ("二手机库存占比", 0.20, PCT, ""),
    ("配件库存占比", 0.10, PCT, ""),
    ("首批库存总预算", 35000, RM0, "建议不超过启动预算的50-60%，分批采购，见文件3.6.4"),
])

r = add_assumption_block(r, "销售规模假设（开业首月，可调整，月2起按增长率递增）", [
    ("月均手机销售台数", 40, "0", "0→1阶段保守假设，建议用首月实际数据校正"),
    ("平均每台手机售价", 1400, RM0, "文件产品定位 RM500-3000 区间的中位数附近"),
    ("配件月均订单数", 150, "0", ""),
    ("配件平均客单价", 45, RM0, ""),
    ("月销量环比增长率", 0.03, PCT, "假设口碑与线上引流带动缓慢增长"),
])

r = add_assumption_block(r, "支付渠道占比与手续费假设 (文件 5.6)", [
    ("现金转账占比", 0.45, PCT, "无手续费"),
    ("DuitNow QR占比", 0.20, PCT, ""),
    ("信用卡占比", 0.15, PCT, ""),
    ("PayLater占比", 0.20, PCT, "文件区间3-6%取中间偏高"),
    ("DuitNow QR费率", 0.010, PCT, ""),
    ("信用卡费率", 0.015, PCT, ""),
    ("PayLater费率", 0.045, PCT, ""),
])

r = add_assumption_block(r, "人员假设 (文件 2.14)", [
    ("员工人数", 1, "0", "开业初期1名销售员工"),
    ("员工基本工资", 2000, RM0, ""),
    ("雇主EPF SOCSO EIS比例", 0.13, PCT, "约13%"),
    ("销售佣金占毛利比例", 0.03, PCT, "假设值，需与员工另行约定"),
])

ws.sheet_view.showGridLines = False


def A(name):
    """Reference an assumption cell by its label, e.g. A('新机毛利率') -> "'关键假设'!C15" """
    return f"'{ASSUMP_SHEET}'!{AC[name]}"


# =====================================================================
# SHEET 2: CAPEX 开店资本支出
# =====================================================================
ws = wb.create_sheet("CAPEX 开店资本支出")
set_col_widths(ws, [3, 34, 14, 14, 14, 40])
style_title(ws, "A1:F1", "CAPEX · 开店一次性资本支出（对应文件第3章 Checklist）")

subtotal_cells = []
cash_reserve_capex_row = None


def capex_section(start_row, title, items):
    """items: (name, qty, unit_cost, note); returns (next_row, subtotal_coord)"""
    style_section(ws, start_row, 2, 5, title)
    r = start_row + 1
    style_header_row(ws, r, 2, ["项目", "数量", "单价 (RM)", "小计 (RM)", "备注"])
    r += 1
    first_data_row = r
    for name, qty, unit_cost, note in items:
        ws.cell(row=r, column=2, value=name).border = border
        input_cell(ws, f"C{r}", qty, "0")
        input_cell(ws, f"D{r}", unit_cost, RM0)
        f = ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
        f.number_format = RM
        f.border = border
        n = ws.cell(row=r, column=6, value=note)
        n.font = note_font
        n.alignment = Alignment(wrap_text=True)
        r += 1
    last_data_row = r - 1
    ws.cell(row=r, column=2, value="小计").font = total_font
    ws.cell(row=r, column=2).fill = total_fill
    for col in (3, 4):
        ws.cell(row=r, column=col).fill = total_fill
    tot = ws.cell(row=r, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})")
    tot.number_format = RM
    tot.font = total_font
    tot.fill = total_fill
    return r + 2, tot.coordinate


def capex_formula_section(start_row, title, rows):
    """rows: (name, formula, note); returns (next_row, subtotal_coord)"""
    style_section(ws, start_row, 2, 5, title)
    rr = start_row + 1
    style_header_row(ws, rr, 2, ["项目", "数量", "单价 (RM)", "小计 (RM)", "备注"])
    rr += 1
    first = rr
    for name, formula, note in rows:
        ws.cell(row=rr, column=2, value=name).border = border
        ws.cell(row=rr, column=3).border = border
        ws.cell(row=rr, column=4).border = border
        c = ws.cell(row=rr, column=5, value=formula)
        c.number_format = RM
        c.border = border
        n = ws.cell(row=rr, column=6, value=note)
        n.font = note_font
        n.alignment = Alignment(wrap_text=True)
        rr += 1
    last = rr - 1
    ws.cell(row=rr, column=2, value="小计").font = total_font
    ws.cell(row=rr, column=2).fill = total_fill
    ws.cell(row=rr, column=3).fill = total_fill
    ws.cell(row=rr, column=4).fill = total_fill
    tot = ws.cell(row=rr, column=5, value=f"=SUM(E{first}:E{last})")
    tot.number_format = RM
    tot.font = total_font
    tot.fill = total_fill
    return rr + 2, tot.coordinate


r = 3
r, t = capex_section(r, "1. 商场租赁 一次性费用 (文件 3.4)", [
    ("租金押金（一般2-3个月租金）", 1, 6000, "待Level B报价确认，见文件7.6商场询价问题清单"),
    ("水电押金", 1, 500, ""),
    ("装修押金", 1, 1000, ""),
    ("招牌申请费/图纸审批费", 1, 800, ""),
    ("律师费/印花税（租约）", 1, 500, ""),
])
subtotal_cells.append(("商场租赁一次性费用", t))

r, t = capex_section(r, "2. 装修工程 (文件 3.8)", [
    ("展示柜（可移动/可重复使用款）", 3, 800, "文件强调优先可移动、未来可搬走"),
    ("柜台/收银台", 1, 1200, ""),
    ("招牌制作与安装", 1, 2500, ""),
    ("灯光工程", 1, 1000, ""),
    ("电源/网络布线", 1, 800, ""),
    ("基础油漆与地板整理", 1, 1500, ""),
    ("装修其他杂项", 1, 1000, ""),
])
subtotal_cells.append(("装修工程", t))

r, t = capex_section(r, "3. 店铺设备 (文件 3.9)", [
    ("CCTV 系统", 1, 900, ""),
    ("电脑/Laptop", 1, 2200, ""),
    ("打印机", 1, 400, ""),
    ("Cash Drawer 收银箱", 1, 250, ""),
    ("Wi-Fi 路由器", 1, 200, ""),
    ("POS/条码扫描器（如采用）", 1, 600, "文件4.4：第一阶段用Google Sheets，POS延后，此项可为0"),
    ("手机防盗展示器/小型保险箱", 1, 500, ""),
    ("灭火器/UPS等安全设备", 1, 300, ""),
    ("包装袋/收据本/标签机等开业物料", 1, 500, ""),
])
subtotal_cells.append(("店铺设备", t))

r, t = capex_section(r, "4. 支付与系统开办费 (文件 3.10, 5.6)", [
    ("DuitNow QR 申请/开通费", 1, 0, "通常免费，待Level B确认"),
    ("信用卡终端机 Setup Fee", 1, 300, "待HitPay/Revenue Monster报价，见文件5.8"),
    ("PayLater 平台 Setup Fee", 1, 300, ""),
    ("WhatsApp Business 设置（人力工时，非现金支出可填0）", 1, 0, ""),
    ("自建网站建置费", 1, 3000, "文件3.11.2：如短期内无法完成付款功能，可先用WhatsApp接单"),
])
subtotal_cells.append(("支付与系统开办费", t))

r, t = capex_section(r, "5. 公司与合规开办费 (呼应待完成的第9章 法律与文件)", [
    ("SSM 年费/文件（如需更新）", 1, 300, "文件3.2：SSM注册已完成，此处为后续年度维护提醒"),
    ("商业执照 Business Premise License", 1, 500, "地方政府牌照，按店铺所在地议会收费"),
    ("招牌执照 Signboard License", 1, 300, ""),
    ("初期法律/合约审阅费（租约、合伙协议）", 1, 1500, "对应文件2.15合伙人书面协议、9章法律文件缺口"),
    ("保险首年保费（如决定购买）", 1, 0, "文件3.14：暂不立即购买，先研究；此处先留0，谈妥后再填"),
])
subtotal_cells.append(("公司与合规开办费", t))

r, t = capex_formula_section(r, "6. 首批库存 (文件 3.6 / Decision Log #008，自动按 70/20/10 比例计算)", [
    ("新机首批库存", f"={A('首批库存总预算')}*{A('新机库存占比')}", "计算式：首批库存总预算 × 新机占比（关键假设表）"),
    ("二手机首批库存", f"={A('首批库存总预算')}*{A('二手机库存占比')}", "计算式：首批库存总预算 × 二手机占比"),
    ("配件首批库存", f"={A('首批库存总预算')}*{A('配件库存占比')}", "计算式：首批库存总预算 × 配件占比"),
])
subtotal_cells.append(("首批库存", t))

r, t = capex_section(r, "7. 现金流保护储备 (文件 3.5 现金流保护原则 / Decision Log #018)", [
    ("应急/营运现金储备（自动引用「OPEX」表建议值，见备注）", 1, 0, "开业当天不得把预算全部用完；此小计会在OPEX表生成后自动改写为公式"),
])
cash_reserve_capex_row = r - 2  # row of the single item above
subtotal_cells.append(("现金流保护储备", t))

gt_row = r + 1
style_section(ws, gt_row, 2, 5, "CAPEX 总计")
gt_row += 1
ws.cell(row=gt_row, column=2, value="CAPEX 总计").font = Font(bold=True, size=12)
formula_parts = "+".join(coord for _, coord in subtotal_cells)
grand = ws.cell(row=gt_row, column=5, value=f"={formula_parts}")
grand.number_format = RM
grand.font = Font(bold=True, size=12)
grand.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws.cell(row=gt_row, column=2).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

gt_row += 1
ws.cell(row=gt_row, column=2, value="对比启动预算下限").font = Font(italic=True)
diff = ws.cell(row=gt_row, column=5, value=f"={A('启动预算下限')}-{grand.coordinate}")
diff.number_format = RM
diff.font = Font(italic=True)

gt_row += 1
ws.cell(row=gt_row, column=2, value="对比启动预算上限").font = Font(italic=True)
diff2 = ws.cell(row=gt_row, column=5, value=f"={A('启动预算上限')}-{grand.coordinate}")
diff2.number_format = RM
diff2.font = Font(italic=True)

CAPEX_TOTAL_CELL = grand.coordinate
CAPEX_SHEET = "CAPEX 开店资本支出"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B3"

# =====================================================================
# SHEET 3: OPEX 每月营运支出
# =====================================================================
ws = wb.create_sheet("OPEX 每月营运支出")
set_col_widths(ws, [3, 34, 14, 40])
style_title(ws, "A1:D1", "OPEX · 每月营运固定/变动支出（对应文件2.16核心成本结构）")

opex_subtotals = []


def opex_section(start_row, title, items):
    """items: (name, value_or_formula, note, is_formula)"""
    style_section(ws, start_row, 2, 3, title)
    r = start_row + 1
    style_header_row(ws, r, 2, ["项目", "每月金额 (RM)", "备注"])
    r += 1
    first = r
    for name, val, note, is_formula in items:
        ws.cell(row=r, column=2, value=name).border = border
        c = ws.cell(row=r, column=3, value=val)
        c.number_format = RM
        c.border = border
        if not is_formula:
            c.fill = input_fill
        n = ws.cell(row=r, column=4, value=note)
        n.font = note_font
        n.alignment = Alignment(wrap_text=True)
        r += 1
    last = r - 1
    ws.cell(row=r, column=2, value="小计").font = total_font
    ws.cell(row=r, column=2).fill = total_fill
    ws.cell(row=r, column=3).fill = total_fill
    tot = ws.cell(row=r, column=3, value=f"=SUM(C{first}:C{last})")
    tot.number_format = RM
    tot.font = total_font
    tot.fill = total_fill
    return r + 2, tot.coordinate


r = 3
r, t = opex_section(r, "1. 场地相关 (文件 3.4)", [
    ("商场基本租金", 3500, "待商场Level B报价，见文件7.9联系清单", False),
    ("管理费/推广费", 400, "", False),
    ("水电费", 500, "", False),
    ("冷气费（如另计）", 200, "", False),
    ("网络/通讯费", 150, "", False),
    ("垃圾处理费/其他杂费", 100, "", False),
])
opex_subtotals.append(("场地相关", t))

r, t = opex_section(r, "2. 人员成本 (文件 2.16 / 自动引用关键假设)", [
    ("员工基本工资", f"={A('员工基本工资')}", "引用关键假设表·员工基本工资", True),
    ("EPF/SOCSO/EIS 雇主部分", f"={A('员工基本工资')}*{A('雇主EPF SOCSO EIS比例')}", "计算式：工资 × 雇主负担比例", True),
    ("销售佣金（此处保留0，避免与「12个月现金流预测」表按毛利动态计算的佣金重复扣除）", 0, "佣金已在「12个月现金流预测」表按当月毛利×佣金比例自动计算，此处若填数字会被重复扣除两次", False),
])
opex_subtotals.append(("人员成本", t))

r, t = opex_section(r, "3. 支付渠道手续费（变动成本，随营收自动计算）", [
    ("说明：手续费为变动成本，已在「12个月现金流预测」按渠道占比与费率自动计算，此处填0避免重复计算", 0, "如需查看月度手续费金额，见「12个月现金流预测」表", False),
])
opex_subtotals.append(("支付渠道手续费", t))

r, t = opex_section(r, "4. 营销与获客 (文件 2.13, 3.15)", [
    ("Facebook 广告", 500, "", False),
    ("TikTok 内容/广告", 300, "", False),
    ("开业/促销物料（首3个月后转为常态小额预算）", 200, "", False),
])
opex_subtotals.append(("营销与获客", t))

r, t = opex_section(r, "5. 系统与订阅 (文件 4.3-4.4, 3.11)", [
    ("Google Sheets（官方库存/记录方案，免费-约RM300/年按月折算）", 25, "文件4.4：Decision Log #007已定为免费/低成本方案", False),
    ("网站维护/域名/主机", 100, "", False),
    ("WhatsApp Business（免费，如用第三方工具另计）", 0, "", False),
])
opex_subtotals.append(("系统与订阅", t))

r, t = opex_section(r, "6. 售后与风险准备 (文件 2.12, 3.13, 3.14 / 呼应第9章法律缺口)", [
    ("维修/售后处理准备金", 300, "对应外部维修商合作费用及非原厂保修处理", False),
    ("保险分摊（如决定购买，见文件3.14）", 0, "暂不购买，先研究；决定后按年保费/12填入", False),
    ("坏账/退换损耗准备（约营收0.5%，估算值）", 150, "", False),
])
opex_subtotals.append(("售后与风险准备", t))

r, t = opex_section(r, "7. 其他行政 (呼应待完成第9章法律与文件)", [
    ("会计/记账服务", 250, "", False),
    ("银行手续费/其他行政杂费", 50, "", False),
    ("LHDN e-Invoice / SST 合规预留（未达门槛前可填0）", 0, "营业额达SST门槛(RM500k)或e-Invoice强制阶段前预留", False),
])
opex_subtotals.append(("其他行政", t))

gt_row = r + 1
style_section(ws, gt_row, 2, 3, "OPEX 总计")
gt_row += 1
ws.cell(row=gt_row, column=2, value="OPEX 总计").font = Font(bold=True, size=12)
formula_parts = "+".join(coord for _, coord in opex_subtotals)
opex_grand = ws.cell(row=gt_row, column=3, value=f"={formula_parts}")
opex_grand.number_format = RM
opex_grand.font = Font(bold=True, size=12)
opex_grand.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws.cell(row=gt_row, column=2).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

gt_row += 1
ws.cell(row=gt_row, column=2, value="建议现金储备（2.5个月固定成本，文件3.5要求2-3个月）").font = Font(italic=True)
reserve = ws.cell(row=gt_row, column=3, value=f"={opex_grand.coordinate}*2.5")
reserve.number_format = RM
reserve.font = Font(italic=True)

OPEX_TOTAL_CELL = opex_grand.coordinate
OPEX_SHEET = "OPEX 每月营运支出"
CASH_RESERVE_CELL = reserve.coordinate
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B3"

# now wire CAPEX's cash-reserve line to this OPEX suggestion
ws_capex = wb[CAPEX_SHEET]
e_cell = ws_capex.cell(row=cash_reserve_capex_row, column=5)
e_cell.value = f"='{OPEX_SHEET}'!{CASH_RESERVE_CELL}"
e_cell.number_format = RM

# =====================================================================
# SHEET 4: 12个月现金流与损益预测
# =====================================================================
ws = wb.create_sheet("12个月现金流预测")
set_col_widths(ws, [30] + [12] * 12 + [14])
style_title(ws, "A1:N1", "12个月现金流与损益预测（自动引用「关键假设」「OPEX」，月2起按增长率递增）")

months = [f"第{i}月" for i in range(1, 13)]
style_header_row(ws, 3, 1, ["项目"] + months + ["合计"])


def col_letter_for_month(i):  # i=0..11 -> columns B..M
    return get_column_letter(2 + i)


row = 4
ws.cell(row=row, column=1, value="手机销量（台，月2起按增长率递增）")
for i in range(12):
    if i == 0:
        c = ws.cell(row=row, column=2, value=f"={A('月均手机销售台数')}")
    else:
        prev = col_letter_for_month(i - 1)
        c = ws.cell(row=row, column=2 + i, value=f"={prev}{row}*(1+{A('月销量环比增长率')})")
    c.number_format = "0.0"
qty_row = row

row += 1
ws.cell(row=row, column=1, value="平均单台售价 (RM)")
for i in range(12):
    ws.cell(row=row, column=2 + i, value=f"={A('平均每台手机售价')}").number_format = RM
price_row = row

row += 1
ws.cell(row=row, column=1, value="手机销售额 (RM)")
for i in range(12):
    col = col_letter_for_month(i)
    ws.cell(row=row, column=2 + i, value=f"={col}{qty_row}*{col}{price_row}").number_format = RM
phone_rev_row = row

row += 1
ws.cell(row=row, column=1, value="配件订单数")
for i in range(12):
    if i == 0:
        c = ws.cell(row=row, column=2, value=f"={A('配件月均订单数')}")
    else:
        prev = col_letter_for_month(i - 1)
        c = ws.cell(row=row, column=2 + i, value=f"={prev}{row}*(1+{A('月销量环比增长率')})")
    c.number_format = "0.0"
acc_qty_row = row

row += 1
ws.cell(row=row, column=1, value="配件平均客单价 (RM)")
for i in range(12):
    ws.cell(row=row, column=2 + i, value=f"={A('配件平均客单价')}").number_format = RM
acc_price_row = row

row += 1
ws.cell(row=row, column=1, value="配件销售额 (RM)")
for i in range(12):
    col = col_letter_for_month(i)
    ws.cell(row=row, column=2 + i, value=f"={col}{acc_qty_row}*{col}{acc_price_row}").number_format = RM
acc_rev_row = row

row += 1
ws.cell(row=row, column=1, value="总营收 (RM)").font = total_font
for i in range(12):
    col = col_letter_for_month(i)
    c = ws.cell(row=row, column=2 + i, value=f"={col}{phone_rev_row}+{col}{acc_rev_row}")
    c.number_format = RM
    c.font = total_font
    c.fill = total_fill
total_rev_row = row

row += 1
ws.cell(row=row, column=1, value="估算毛利 (RM)  新机/二手机/配件加权")
new_share = A('新机收入占比')
used_share = A('二手机收入占比')
new_margin = A('新机毛利率')
used_margin = A('二手机毛利率')
acc_margin = A('配件毛利率')
for i in range(12):
    col = col_letter_for_month(i)
    phone_margin_formula = f"({col}{phone_rev_row}*(({new_share}/({new_share}+{used_share}))*{new_margin}+({used_share}/({new_share}+{used_share}))*{used_margin}))"
    acc_margin_formula = f"({col}{acc_rev_row}*{acc_margin})"
    c = ws.cell(row=row, column=2 + i, value=f"={phone_margin_formula}+{acc_margin_formula}")
    c.number_format = RM
gross_margin_row = row

row += 1
ws.cell(row=row, column=1, value="支付手续费 (RM)")
cash_share = A('现金转账占比')
duitnow_share = A('DuitNow QR占比')
card_share = A('信用卡占比')
paylater_share = A('PayLater占比')
duitnow_fee = A('DuitNow QR费率')
card_fee = A('信用卡费率')
paylater_fee = A('PayLater费率')
for i in range(12):
    col = col_letter_for_month(i)
    fee = f"={col}{total_rev_row}*({duitnow_share}*{duitnow_fee}+{card_share}*{card_fee}+{paylater_share}*{paylater_fee})"
    ws.cell(row=row, column=2 + i, value=fee).number_format = RM
fee_row = row

row += 1
ws.cell(row=row, column=1, value="每月固定营运支出 OPEX (RM)")
for i in range(12):
    ws.cell(row=row, column=2 + i, value=f"='{OPEX_SHEET}'!{OPEX_TOTAL_CELL}").number_format = RM
opex_row = row

row += 1
ws.cell(row=row, column=1, value="销售佣金 (毛利的比例)")
for i in range(12):
    col = col_letter_for_month(i)
    ws.cell(row=row, column=2 + i, value=f"={col}{gross_margin_row}*{A('销售佣金占毛利比例')}").number_format = RM
commission_row = row

row += 1
ws.cell(row=row, column=1, value="净利润 (RM) = 毛利 − 手续费 − OPEX − 佣金").font = total_font
for i in range(12):
    col = col_letter_for_month(i)
    c = ws.cell(row=row, column=2 + i, value=f"={col}{gross_margin_row}-{col}{fee_row}-{col}{opex_row}-{col}{commission_row}")
    c.number_format = RM
    c.font = total_font
    c.fill = total_fill
net_profit_row = row

row += 1
ws.cell(row=row, column=1, value="累计净现金流 (RM，起点为0，未计CAPEX)")
for i in range(12):
    col = col_letter_for_month(i)
    if i == 0:
        ws.cell(row=row, column=2 + i, value=f"={col}{net_profit_row}").number_format = RM
    else:
        prev = col_letter_for_month(i - 1)
        ws.cell(row=row, column=2 + i, value=f"={prev}{row}+{col}{net_profit_row}").number_format = RM
cum_row = row

# totals column (N)
n_col = 14
for r_ in [qty_row, phone_rev_row, acc_qty_row, acc_rev_row, total_rev_row, gross_margin_row, fee_row, opex_row, commission_row, net_profit_row]:
    c = ws.cell(row=r_, column=n_col, value=f"=SUM(B{r_}:M{r_})")
    c.number_format = "0.0" if r_ in (qty_row, acc_qty_row) else RM
    c.font = total_font

ws.cell(row=3, column=n_col, value="12月合计").font = header_font
ws.cell(row=3, column=n_col).fill = header_fill
ws.cell(row=3, column=n_col).border = border

row += 2
ws.cell(row=row, column=1, value="Decision Log #024 验证标准：连续3个月净利润>0 且累计净现金流为正").font = Font(italic=True, size=9, color="666666")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)

red_font = Font(color=CRIT)
green_font = Font(color=GOOD)
ws.conditional_formatting.add(f"B{net_profit_row}:M{net_profit_row}", CellIsRule(operator='lessThan', formula=['0'], font=red_font))
ws.conditional_formatting.add(f"B{net_profit_row}:M{net_profit_row}", CellIsRule(operator='greaterThanOrEqual', formula=['0'], font=green_font))
ws.conditional_formatting.add(f"B{cum_row}:N{cum_row}", CellIsRule(operator='lessThan', formula=['0'], font=red_font))
ws.conditional_formatting.add(f"B{cum_row}:N{cum_row}", CellIsRule(operator='greaterThanOrEqual', formula=['0'], font=green_font))

ws.sheet_view.showGridLines = False
ws.freeze_panes = "B4"
FORECAST_SHEET = "12个月现金流预测"
FORECAST_TOTAL_REV_CELL = f"N{total_rev_row}"
FORECAST_TOTAL_NETPROFIT_CELL = f"N{net_profit_row}"
FORECAST_LAST_CUM_CELL = f"M{cum_row}"

# =====================================================================
# SHEET 5: 总览 Summary Dashboard
# =====================================================================
ws = wb.create_sheet("总览")
wb.move_sheet("总览", offset=-5)  # place right after README, before 关键假设
set_col_widths(ws, [4, 34, 20, 40])
style_title(ws, "A1:D1", "SUN SMART · CAPEX/OPEX 总览")

ws["B3"] = "货币单位：马来西亚令吉 (RM)　|　所有数字最终来源：「关键假设」表"
ws["B3"].font = note_font

r = 5
style_section(ws, r, 2, 2, "一次性资本支出 CAPEX")
r += 1
ws.cell(row=r, column=2, value="CAPEX 总计（含现金储备）")
c = ws.cell(row=r, column=3, value=f"='{CAPEX_SHEET}'!{CAPEX_TOTAL_CELL}")
c.number_format = RM
c.font = Font(bold=True, size=13)
capex_total_ref = f"'总览'!{c.coordinate}"
r += 2

style_section(ws, r, 2, 2, "每月营运支出 OPEX")
r += 1
ws.cell(row=r, column=2, value="每月 OPEX 总计")
c = ws.cell(row=r, column=3, value=f"='{OPEX_SHEET}'!{OPEX_TOTAL_CELL}")
c.number_format = RM
c.font = Font(bold=True, size=13)
r += 1
ws.cell(row=r, column=2, value="建议现金储备（2.5个月）")
c = ws.cell(row=r, column=3, value=f"='{OPEX_SHEET}'!{CASH_RESERVE_CELL}")
c.number_format = RM
r += 2

style_section(ws, r, 2, 2, "总启动资金需求")
r += 1
ws.cell(row=r, column=2, value="CAPEX 总计（已含现金储备）").font = Font(bold=True)
c = ws.cell(row=r, column=3, value=f"='{CAPEX_SHEET}'!{CAPEX_TOTAL_CELL}")
c.number_format = RM
c.font = Font(bold=True, size=13)
total_need_cell = c.coordinate
r += 1
ws.cell(row=r, column=2, value="对比 RM50,000-100,000 预算区间")
c = ws.cell(
    row=r, column=3,
    value=f"=IF({total_need_cell}<={A('启动预算下限')},\"低于下限，资金充裕\",IF({total_need_cell}<={A('启动预算上限')},\"落在预算区间内\",\"超出预算上限，需调整假设或增资\"))"
)
r += 3

style_section(ws, r, 2, 2, "12个月预测摘要")
r += 1
ws.cell(row=r, column=2, value="12个月总营收")
c = ws.cell(row=r, column=3, value=f"='{FORECAST_SHEET}'!{FORECAST_TOTAL_REV_CELL}")
c.number_format = RM
r += 1
ws.cell(row=r, column=2, value="12个月净利润合计")
c = ws.cell(row=r, column=3, value=f"='{FORECAST_SHEET}'!{FORECAST_TOTAL_NETPROFIT_CELL}")
c.number_format = RM
r += 1
ws.cell(row=r, column=2, value="第12月末累计净现金流")
c = ws.cell(row=r, column=3, value=f"='{FORECAST_SHEET}'!{FORECAST_LAST_CUM_CELL}")
c.number_format = RM
r += 2

note = ws.cell(row=r, column=2, value="提示：本表所有黄色输入格集中在「关键假设」「CAPEX」「OPEX」三张表，改完后回到本页查看总览是否仍落在预算区间内。")
note.font = note_font
note.alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
ws.row_dimensions[r].height = 30

ws.sheet_view.showGridLines = False

# Open on the "总览" (Summary) tab by default so the CAPEX/OPEX totals are the
# first thing visible, instead of the text-only "使用说明" instructions tab.
wb.active = wb.sheetnames.index("总览")

OUTPUT_PATH = "SUN_SMART_CAPEX_OPEX.xlsx"
wb.save(OUTPUT_PATH)

from bake_cache import bake
baked = bake(OUTPUT_PATH)
print(f"Saved {OUTPUT_PATH} (baked {baked} cached formula value(s) for non-recalculating viewers)")
print("Assumption cell map:")
for k, v in AC.items():
    print(f"  {k}: {v}")

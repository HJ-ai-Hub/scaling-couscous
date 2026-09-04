"""Static audit of the generated workbook, independent of LibreOffice.

Catches the failure modes that matter here:
  1. References to sheets that do not exist.
  2. Functions LibreOffice/Excel-2007 cannot evaluate without a prefix.
  3. Off-by-one references — a formula pointing at a cell that is empty or holds
     a text label instead of a number. This is the silent killer: the file
     recalculates cleanly and reports the wrong answer.
  4. Self-referencing formulas (circular).
"""

import re
import sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

PATH = sys.argv[1] if len(sys.argv) > 1 else "Orizuru_CAPEX_OPEX_Model.xlsx"

BANNED = {"XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE",
          "TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS"}

wb = load_workbook(PATH)
sheets = set(wb.sheetnames)

REF = re.compile(
    r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))?!?\$?([A-Z]{1,3})\$?(\d+)(?::\$?([A-Z]{1,3})\$?(\d+))?"
)
SHEETREF = re.compile(r"(?:'([^']+)'|(\b[A-Za-z][A-Za-z0-9_]*))!")
FUNC = re.compile(r"\b([A-Z][A-Z0-9_.]{1,20})\s*\(")

problems = []
text_refs = set()
n_formulas = 0
cell_ref = re.compile(r"(?:(?:'([^']+)'|([A-Za-z][A-Za-z0-9_]*))!)?\$?([A-Z]{1,3})\$?(\d+)")


def is_blankish(ws, col, row):
    v = ws.cell(row, col).value
    return v is None or (isinstance(v, str) and not v.startswith("="))


for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            f = c.value
            if not isinstance(f, str) or not f.startswith("="):
                continue
            n_formulas += 1

            # 1. sheet names
            for q, u in SHEETREF.findall(f):
                name = q or u
                if name not in sheets:
                    problems.append(f"{ws.title}!{c.coordinate}: unknown sheet '{name}'  ->  {f}")

            # 2. banned functions
            for fn in FUNC.findall(f):
                base = fn.split(".")[-1]
                if base in BANNED and "_xlfn." not in f:
                    problems.append(f"{ws.title}!{c.coordinate}: banned/unprefixed function {fn}  ->  {f}")

            # 3. single-cell refs pointing at blank or text cells
            body = f[1:]
            # strip quoted strings so text literals do not parse as refs
            body_nostr = re.sub(r'"[^"]*"', "", body)
            for m in cell_ref.finditer(body_nostr):
                q, u, col_l, row_n = m.groups()
                # skip if part of a range (followed by ':')
                end = m.end()
                if end < len(body_nostr) and body_nostr[end] == ":":
                    continue
                start = m.start()
                if start > 0 and body_nostr[start - 1] == ":":
                    continue
                tname = q or u
                target = wb[tname] if tname in sheets else (None if tname else ws)
                if target is None:
                    continue
                col_i = 0
                for ch in col_l:
                    col_i = col_i * 26 + (ord(ch) - 64)
                if target is ws and col_i == c.column and int(row_n) == c.row:
                    problems.append(f"{ws.title}!{c.coordinate}: self-reference (circular)  ->  {f}")
                    continue
                v = target.cell(int(row_n), col_i).value
                if v is None:
                    problems.append(
                        f"{ws.title}!{c.coordinate}: refers to EMPTY cell "
                        f"{tname or ws.title}!{col_l}{row_n}  ->  {f}")
                elif isinstance(v, str) and not v.startswith("="):
                    # only a defect if the text is being used arithmetically
                    ref_txt = m.group(0)
                    i0, i1 = m.start(), m.end()
                    before = body_nostr[max(0, i0 - 1):i0]
                    after = body_nostr[i1:i1 + 1]
                    if before in "+-*/^" or after in "+-*/^":
                        problems.append(
                            f"{ws.title}!{c.coordinate}: arithmetic on TEXT cell "
                            f"{tname or ws.title}!{col_l}{row_n} = {v!r}  ->  {f}")
                    else:
                        text_refs.add(f"{tname or ws.title}!{col_l}{row_n} = {v!r}")

print(f"file            : {PATH}")
print(f"sheets          : {len(sheets)}")
print(f"formulas checked: {n_formulas}")
print(f"text-switch refs: {len(text_refs)} (expected - scenario/option selectors)")
for t in sorted(text_refs):
    print("  .", t)
print(f"problems        : {len(problems)}")
for p in problems:
    print("  -", p)
sys.exit(1 if problems else 0)

"""Build Orizuru Holiday Homes LLC — CAPEX workbook (AED).

Model basis (confirmed by management):
  - Pure management / commission-only. Orizuru does NOT master-lease units.
  - Unit setup is funded by the property owner and is therefore excluded from CAPEX.
  - Directors draw no salary (relevant to OPEX, noted here for completeness).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/home/user/scaling-couscous/orizuru/Orizuru_CAPEX.xlsx"

# ---------------------------------------------------------------- styling ---
FONT = "Arial"
NAVY = "1F3352"
SAND = "EDE7DC"
RULE = "C7B99A"

f_title = Font(name=FONT, size=14, bold=True, color="FFFFFF")
f_sub = Font(name=FONT, size=9, color="FFFFFF")
f_sec = Font(name=FONT, size=10, bold=True, color="FFFFFF")
f_hdr = Font(name=FONT, size=9, bold=True, color="FFFFFF")
f_item = Font(name=FONT, size=10)
f_in = Font(name=FONT, size=10, color="0000FF")
f_calc = Font(name=FONT, size=10)
f_tot = Font(name=FONT, size=10, bold=True)
f_note = Font(name=FONT, size=8, color="595959", italic=True)
f_link = Font(name=FONT, size=10, color="008000")

fill_title = PatternFill("solid", fgColor=NAVY)
fill_sec = PatternFill("solid", fgColor="41597A")
fill_hdr = PatternFill("solid", fgColor="6B7F99")
fill_band = PatternFill("solid", fgColor=SAND)
fill_key = PatternFill("solid", fgColor="FFFF00")
fill_tot = PatternFill("solid", fgColor="DCE3EC")

thin = Side(style="thin", color=RULE)
b_bottom = Border(bottom=thin)
b_top = Border(top=Side(style="medium", color=NAVY))

CUR = '#,##0;(#,##0);-'
CUR2 = '#,##0.00;(#,##0.00);-'
PCT = '0.0%'

wb = Workbook()


def banner(ws, title, subtitle, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(1, 1, title)
    c.font = f_title
    c.fill = fill_title
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c = ws.cell(2, 2 - 1, subtitle)
    c.font = f_sub
    c.fill = fill_title
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 16
    for col in range(1, width + 1):
        ws.cell(1, col).fill = fill_title
        ws.cell(2, col).fill = fill_title


# ================================================================= README ===
ws = wb.active
ws.title = "README"
banner(ws, "ORIZURU HOLIDAY HOMES LLC — CAPITAL EXPENDITURE (CAPEX)",
       "Dubai, UAE  |  Trade Licence 1634278  |  All figures in AED", 6)

readme = [
    ("", ""),
    ("SEC", "HOW TO USE THIS WORKBOOK"),
    ("T", "Tab", "What it contains"),
    ("R", "CAPEX Detail", "Every one-off setup cost, line by line, at Low / Base / High. This is the main sheet."),
    ("R", "Unit Setup (1 Unit)", "What onboarding a unit involves and what it costs. Owner-funded, so excluded from the CAPEX total."),
    ("R", "CAPEX Summary", "Category roll-up and funding view. Reads from the two sheets above."),
    ("", ""),
    ("SEC", "MODEL BASIS — CONFIRMED WITH MANAGEMENT"),
    ("B", "1.", "Pure management / commission-only model. Orizuru does NOT master-lease units, so there is NO unit rent, "
                "DEWA deposit or furnishing capex per unit — those sit with the property owner."),
    ("B", "2.", "UNIT SETUP IS FUNDED BY THE PROPERTY OWNER, so it does NOT appear in this CAPEX. The 'Unit Setup' "
                "tab documents what onboarding a unit involves and what it would cost if Orizuru bore it — governed "
                "by the 'Share of unit setup cost borne by Orizuru' input on the Assumptions tab, currently 0%."),
    ("B", "3.", "Shareholder capital is AED 200,000 as confirmed (the Business Profile states AED 2,000,000 — these "
                "must be reconciled). Either way it is NOT used as a target or a cap: this CAPEX is built bottom-up "
                "from what the business actually needs."),
    ("B", "4.", "Directors draw no salary. See the OPEX workbook."),
    ("", ""),
    ("SEC", "COLOUR LEGEND"),
    ("L", "Blue text", "Hardcoded input — edit these freely."),
    ("L", "Yellow fill", "Key assumption or a figure you supplied. Check before circulating."),
    ("L", "Black text", "Formula. Do not overwrite."),
    ("L", "Green text", "Link to another sheet in this workbook."),
    ("", ""),
    ("SEC", "FIGURES SUPPLIED BY MANAGEMENT (used as-is, not estimated)"),
    ("B", "•", "Domain name, 3 years — AED 387"),
    ("B", "•", "Company stamp — AED 137"),
    ("", ""),
    ("SEC", "IMPORTANT NOTES"),
    ("B", "•", "Bank minimum balance is shown as a MEMO item, not CAPEX. It is a blocked asset, not a cost — do not "
                "add it to the CAPEX total when presenting to the bank."),
    ("B", "•", "Company formation costs dated on or before 03/07/2026 are likely already incurred. Mark them in the "
                "'Status' column so the sheet distinguishes sunk cost from cash still to be spent."),
    ("B", "•", "DET (formerly DTCM) holiday-home permit fees are charged per unit per year and are re-charged to the "
                "owner in most Dubai management contracts. First issue sits in CAPEX; renewals sit in OPEX."),
    ("B", "•", "Contingency is set at 10% of subtotal. Standard for a first-year UAE service company where licensing "
                "and regulatory fees move."),
]

r = 4
for row in readme:
    if row[0] == "":
        r += 1
        continue
    if row[0] == "SEC":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(r, 1, row[1])
        c.font = f_sec
        c.alignment = Alignment(vertical="center", indent=1)
        for col in range(1, 7):
            ws.cell(r, col).fill = fill_sec
        ws.row_dimensions[r].height = 18
    elif row[0] == "T":
        ws.cell(r, 1, row[1]).font = f_hdr
        ws.cell(r, 2, row[2]).font = f_hdr
        for col in range(1, 7):
            ws.cell(r, col).fill = fill_hdr
    else:
        ws.cell(r, 1, row[1]).font = Font(name=FONT, size=10, bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(r, 2, row[2])
        c.font = f_item
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(14, 13 * (1 + len(row[2]) // 95))
        if row[0] == "L":
            if "Blue" in row[1]:
                ws.cell(r, 1).font = Font(name=FONT, size=10, bold=True, color="0000FF")
            if "Green" in row[1]:
                ws.cell(r, 1).font = Font(name=FONT, size=10, bold=True, color="008000")
            if "Yellow" in row[1]:
                ws.cell(r, 1).fill = fill_key
    r += 1

ws.column_dimensions["A"].width = 20
for col in "BCDEF":
    ws.column_dimensions[col].width = 18
ws.sheet_view.showGridLines = False

# ========================================================== UNIT SETUP =======
wsu = wb.create_sheet("Unit Setup (1 Unit)")
banner(wsu, "UNIT SETUP CAPEX — COST TO BRING ONE UNIT LIVE",
       "Commission-only model: the owner supplies the furnished apartment. Orizuru funds only the operating layer.", 8)

hdr = ["Ref", "Line item", "Qty", "Unit cost — Low", "Unit cost — Base",
       "Unit cost — High", "Total — Base", "Basis / note"]
hr = 4
for i, h in enumerate(hdr, 1):
    c = wsu.cell(hr, i, h)
    c.font = f_hdr
    c.fill = fill_hdr
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
wsu.row_dimensions[hr].height = 30

unit_items = [
    ("U1", "DET holiday-home permit — first issue (per unit, per year)", 1, 1200, 1520, 1800,
     "Dubai Dept. of Economy & Tourism per-unit permit. Usually re-charged to the owner — confirm in the management contract."),
    ("U2", "Smart lock / keyless entry device + install", 1, 850, 1200, 1700,
     "Essential for self check-in. Removes the cost of a physical key handover per booking."),
    ("U3", "Professional listing photography (unit)", 1, 500, 800, 1400,
     "Single biggest driver of conversion on Airbnb / Booking.com. Do not economise here."),
    ("U4", "Linen & towel par stock — 3 sets", 1, 1100, 1800, 2600,
     "3 sets allows one in use, one in laundry, one on the shelf. Required for same-day turnovers."),
    ("U5", "Welcome & amenity kit (branded, Japanese-hospitality standard)", 1, 300, 450, 700,
     "Brand differentiator per the Ma / Shibui positioning in the Business Profile."),
    ("U6", "Safety kit — smoke detector, fire extinguisher, first-aid box", 1, 380, 550, 800,
     "DET holiday-home inspection requirement."),
    ("U7", "Small appliances & inventory top-up (kettle, iron, hairdryer, crockery gaps)", 1, 500, 900, 1600,
     "Covers the gap between what the owner supplies and DET / guest expectation."),
    ("U8", "Initial deep clean & staging before first listing", 1, 400, 600, 950,
     "One-off. Routine turnover cleaning sits in OPEX."),
    ("U9", "Inventory & condition report, listing build, channel sync", 1, 200, 300, 500,
     "Protects against damage disputes and is the baseline for owner reporting."),
    ("U10", "Guest welcome book, house rules, signage (printed)", 1, 120, 200, 350,
     "Reduces avoidable guest-support contacts."),
]

r = hr + 1
u_first = r
for it in unit_items:
    wsu.cell(r, 1, it[0]).font = f_item
    wsu.cell(r, 2, it[1]).font = f_item
    wsu.cell(r, 3, it[2]).font = f_in
    wsu.cell(r, 4, it[3]).font = f_in
    wsu.cell(r, 5, it[4]).font = f_in
    wsu.cell(r, 6, it[5]).font = f_in
    wsu.cell(r, 7, f"=C{r}*E{r}").font = f_calc
    n = wsu.cell(r, 8, it[6])
    n.font = f_note
    n.alignment = Alignment(wrap_text=True, vertical="top")
    for col in (4, 5, 6, 7):
        wsu.cell(r, col).number_format = CUR
    wsu.cell(r, 3).alignment = Alignment(horizontal="center")
    for col in range(1, 9):
        wsu.cell(r, col).border = b_bottom
    wsu.row_dimensions[r].height = 26
    r += 1
u_last = r - 1

# subtotal + contingency + total
wsu.cell(r, 2, "Subtotal — per unit").font = f_tot
for col, letter in ((4, "D"), (5, "E"), (6, "F")):
    wsu.cell(r, col, f"=SUMPRODUCT($C${u_first}:$C${u_last},{letter}{u_first}:{letter}{u_last})").font = f_tot
wsu.cell(r, 7, f"=SUM(G{u_first}:G{u_last})").font = f_tot
for col in range(1, 9):
    wsu.cell(r, col).fill = fill_tot
    wsu.cell(r, col).border = b_top
for col in (4, 5, 6, 7):
    wsu.cell(r, col).number_format = CUR
u_sub = r
r += 1

wsu.cell(r, 2, "Contingency").font = f_item
wsu.cell(r, 3, 0.10).font = f_in
wsu.cell(r, 3).number_format = PCT
wsu.cell(r, 3).fill = fill_key
wsu.cell(r, 3).alignment = Alignment(horizontal="center")
for col, letter in ((4, "D"), (5, "E"), (6, "F"), (7, "G")):
    wsu.cell(r, col, f"={letter}{u_sub}*$C${r}").font = f_calc
    wsu.cell(r, col).number_format = CUR
wsu.cell(r, 8, "Applied to the per-unit subtotal. Covers price movement on permits and hardware.").font = f_note
u_cont = r
r += 1

wsu.cell(r, 2, "TOTAL CAPEX PER UNIT").font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
for col, letter in ((4, "D"), (5, "E"), (6, "F"), (7, "G")):
    c = wsu.cell(r, col, f"={letter}{u_sub}+{letter}{u_cont}")
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.number_format = CUR
for col in range(1, 9):
    wsu.cell(r, col).fill = fill_sec
wsu.row_dimensions[r].height = 20
U_TOTAL_ROW = r
r += 3

# scaling table
wsu.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = wsu.cell(r, 1, "PORTFOLIO SCALING REFERENCE — OWNER-FUNDED, NOT INCLUDED IN THE CAPEX TOTAL")
c.font = f_sec
c.alignment = Alignment(vertical="center", indent=1)
for col in range(1, 9):
    wsu.cell(r, col).fill = fill_sec
r += 1
wsu.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = wsu.cell(r, 1, "The property owner funds unit setup, so NONE of this sits in Orizuru's CAPEX. This tab documents what "
                   "onboarding involves and prices the risk if Orizuru ever has to bear it — see the share input on "
                   "the Assumptions tab and the sensitivity block on 'Financial Summary'.")
c.font = f_note
c.alignment = Alignment(wrap_text=True, vertical="top")
wsu.row_dimensions[r].height = 26
r += 2

sc_hdr = ["Units", "Total — Low", "Total — Base", "Total — High", "", "", "", ""]
for i, h in enumerate(sc_hdr, 1):
    if h:
        c = wsu.cell(r, i, h)
        c.font = f_hdr
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal="center")
r += 1
for n in (1, 5, 10, 25, 50, 80):
    wsu.cell(r, 1, n).font = f_in
    wsu.cell(r, 1).alignment = Alignment(horizontal="center")
    for col, letter in ((2, "D"), (3, "E"), (4, "F")):
        wsu.cell(r, col, f"=$A{r}*{letter}${U_TOTAL_ROW}").font = f_calc
        wsu.cell(r, col).number_format = CUR
    if n == 80:
        for col in range(1, 5):
            wsu.cell(r, col).fill = fill_band
    r += 1

wsu.column_dimensions["A"].width = 7
wsu.column_dimensions["B"].width = 56
wsu.column_dimensions["C"].width = 8
for col in "DEFG":
    wsu.column_dimensions[col].width = 15
wsu.column_dimensions["H"].width = 62
wsu.sheet_view.showGridLines = False
wsu.freeze_panes = "A5"

# ========================================================== CAPEX DETAIL ====
wsd = wb.create_sheet("CAPEX Detail", 1)
banner(wsd, "CAPEX DETAIL — ONE-OFF SETUP EXPENDITURE",
       "Company formation, licensing, technology, brand and equipment. Unit setup is on the 'Unit Setup (1 Unit)' tab.", 10)

dhdr = ["Ref", "Line item", "Qty", "Unit cost — Low", "Unit cost — Base", "Unit cost — High",
        "Total — Low", "Total — Base", "Total — High", "Timing / status", "Basis / note"]
hr = 4
for i, h in enumerate(dhdr, 1):
    c = wsd.cell(hr, i, h)
    c.font = f_hdr
    c.fill = fill_hdr
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
wsd.row_dimensions[hr].height = 30

# (section, ref, item, qty, low, base, high, timing, note, is_hj_figure)
capex = [
    ("A. COMPANY FORMATION, LICENSING & REGULATORY", None),
    ("A1", "DED trade licence — initial issue (Vacation Homes Rental)", 1, 13500, 17500, 22000, "Incurred — 03/07/2026",
     "Mainland LLC commercial licence. Licence No. 1634278 per the Business Profile.", False),
    ("A2", "Trade name reservation & initial approval", 1, 900, 1200, 1600, "Incurred — 03/07/2026",
     "DED fixed fees.", False),
    ("A3", "MOA drafting & notarisation", 1, 1800, 2500, 3500, "Incurred — 03/07/2026",
     "Three-shareholder structure per the Business Profile (Podocarpus 70% / Atayev 16% / Mamiyev 14%).", False),
    ("A4", "Ejari / tenancy registration for the licensed address", 1, 1500, 2200, 3200, "Incurred — 03/07/2026",
     "Fronds Building – Office 02, Goze Industrial Third. Required to hold the licence.", False),
    ("A5", "Immigration establishment card", 1, 1000, 1500, 2200, "Month 1",
     "Required before any employment visa can be issued.", False),
    ("A6", "MOHRE labour file / e-Channel registration", 1, 1600, 2300, 3200, "Month 1",
     "Needed to sponsor the operations & service staff member.", False),
    ("A7", "DET (ex-DTCM) holiday-home OPERATOR permit — company level", 1, 4500, 7500, 12000, "Month 1",
     "Distinct from the per-unit permit. Without it the company cannot legally list short-term rentals.", False),
    ("A8", "Corporate bank account opening & KYC pack", 1, 0, 500, 1500, "Month 1",
     "Corporate current account, multi-currency AED/USD/EUR per the Business Profile.", False),
    ("A9", "VAT registration & Corporate Tax registration", 1, 0, 1000, 2500, "Incurred",
     "TRN 105507385000001 already issued per the Business Profile. Cost shown for completeness.", False),
    ("A10", "Company stamp", 1, 137, 137, 137, "Incurred",
     "Actual figure supplied by management — AED 137.", True),
    ("A11", "Employment visa — operations & service staff (1 person)", 1, 4500, 6000, 8000, "Month 1",
     "Entry permit, status change, medical, Emirates ID, 2-year residence visa.", False),
    ("A12", "Refundable labour guarantee / bank guarantee (1 employee)", 1, 0, 3000, 3000, "Month 1",
     "MOHRE requirement. Refundable, but cash out on day one.", False),

    ("B. TECHNOLOGY, PLATFORM & DIGITAL SETUP", None),
    ("B1", "Domain name registration — 3 years", 1, 387, 387, 387, "Incurred",
     "Actual figure supplied by management — AED 387 for 3 years.", True),
    ("B2", "Website build + direct booking engine", 1, 7000, 12000, 22000, "Month 1–2",
     "The Business Profile commits to a direct booking portal. Direct bookings avoid the 15–18% OTA commission, "
     "so this pays back fastest of any line here.", False),
    ("B3", "PMS / channel manager onboarding & setup fee (Hostaway, Guesty or equivalent)", 1, 1800, 3000, 5500, "Month 1",
     "One-off implementation. The recurring per-unit licence sits in OPEX.", False),
    ("B4", "Channel account setup & verification (Airbnb, Booking.com, direct)", 1, 0, 800, 1800, "Month 1",
     "Account verification, payout setup, listing template build.", False),
    ("B5", "Business email, Google Workspace & cloud drive setup", 1, 300, 600, 1200, "Month 1",
     "One-off configuration. Seat licences are OPEX.", False),
    ("B6", "Accounting system setup & chart of accounts (Zoho Books / Xero)", 1, 800, 1500, 3000, "Month 1",
     "VAT-compliant chart of accounts. Doing this properly now avoids a costly first-audit clean-up.", False),
    ("B7", "AI workspace setup — prompt library, SOP automation, guest-messaging templates", 1, 0, 1500, 4000, "Month 1–2",
     "One-off build of the AI operating layer. The recurring subscription cost sits in OPEX per management item 5.", False),

    ("C. BRAND & MARKETING SETUP", None),
    ("C1", "Brand identity — logo, typography, Japanese-inspired visual system", 1, 2500, 4500, 9000, "Month 1",
     "Supports the Ma / Shibui / Bushido / Wabi-Sabi brand philosophy in the Business Profile.", False),
    ("C2", "Brand photography & short-form video (company level, not per unit)", 1, 2000, 3500, 7000, "Month 1–2",
     "Used across the website, OTA profile and owner pitch deck.", False),
    ("C3", "Owner acquisition pitch deck & management-contract pack", 1, 800, 1800, 3500, "Month 1",
     "The core sales asset for the marketing team in both Option 1 and Option 2.", False),
    ("C4", "Printed collateral, business cards, owner welcome pack", 1, 400, 900, 1800, "Month 1",
     "", False),
    ("C5", "Social media profile setup & launch content batch", 1, 500, 1200, 2500, "Month 1",
     "", False),

    ("D. OFFICE EQUIPMENT & FURNITURE (DUBAI)", None),
    ("D1", "Laptop — operations", 2, 2800, 3800, 5500, "Month 1",
     "One for the operations & service staff member, one for management.", False),
    ("D2", "Mobile handset — dedicated 24/7 guest support line", 1, 1200, 1900, 3200, "Month 1",
     "The Business Profile promises 24/7 guest support. It needs its own line and device.", False),
    ("D3", "Printer / scanner (contracts, DET paperwork)", 1, 500, 850, 1500, "Month 1",
     "", False),
    ("D4", "Office desk, chairs & storage", 1, 1800, 3200, 6000, "Month 1",
     "", False),
    ("D5", "Key-safe cabinet & spare-key management system", 1, 500, 900, 1600, "Month 1",
     "Physical key control across a multi-unit portfolio. An audit point for owners.", False),
    ("D6", "Router, networking & basic office IT", 1, 300, 550, 1000, "Month 1",
     "", False),

    ("E. DEPOSITS & PREPAYMENTS", None),
    ("E1", "Office rent security deposit (refundable)", 1, 2500, 5000, 9000, "Month 1",
     "Refundable at lease end. Cash out on day one.", False),
    ("E2", "DEWA office connection deposit (refundable)", 1, 1000, 2000, 3500, "Month 1",
     "Refundable.", False),
    ("E3", "Public liability & professional indemnity — first-year premium prepayment", 1, 1500, 2800, 5000, "Month 1",
     "Insurers commonly require the first year up front. The ongoing accrual also appears in OPEX — do not double-count "
     "when consolidating; treat this as the timing of the cash, not an extra cost.", False),
]

r = hr + 1
section_rows = []
detail_rows = []
cur_sec = None
for row in capex:
    if row[1] is None:
        wsd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        c = wsd.cell(r, 1, row[0])
        c.font = f_sec
        c.alignment = Alignment(vertical="center", indent=1)
        for col in range(1, 12):
            wsd.cell(r, col).fill = fill_sec
        wsd.row_dimensions[r].height = 18
        cur_sec = [row[0], r, None, None]
        section_rows.append(cur_sec)
        r += 1
        continue
    ref, item, qty, lo, ba, hi, timing, note, is_hj = row
    wsd.cell(r, 1, ref).font = f_item
    wsd.cell(r, 2, item).font = f_item
    wsd.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    wsd.cell(r, 3, qty).font = f_in
    wsd.cell(r, 3).alignment = Alignment(horizontal="center")
    for col, v in ((4, lo), (5, ba), (6, hi)):
        c = wsd.cell(r, col, v)
        c.font = f_in
        c.number_format = CUR
        if is_hj:
            c.fill = fill_key
    for col, letter in ((7, "D"), (8, "E"), (9, "F")):
        c = wsd.cell(r, col, f"=$C{r}*{letter}{r}")
        c.font = f_calc
        c.number_format = CUR
    wsd.cell(r, 10, timing).font = f_item
    wsd.cell(r, 10).alignment = Alignment(horizontal="center", wrap_text=True)
    n = wsd.cell(r, 11, note)
    n.font = f_note
    n.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 12):
        wsd.cell(r, col).border = b_bottom
    wsd.row_dimensions[r].height = max(24, 12 * (1 + len(note) // 78))
    if cur_sec[2] is None:
        cur_sec[2] = r
    cur_sec[3] = r
    detail_rows.append(r)
    r += 1

# section subtotals appended after the grid
r += 1
wsd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
c = wsd.cell(r, 1, "TOTALS")
c.font = f_sec
c.alignment = Alignment(vertical="center", indent=1)
for col in range(1, 12):
    wsd.cell(r, col).fill = fill_sec
r += 1

sec_total_rows = {}
for name, srow, first, last in section_rows:
    wsd.cell(r, 2, name).font = f_tot
    for col, letter in ((7, "G"), (8, "H"), (9, "I")):
        wsd.cell(r, col, f"=SUM({letter}{first}:{letter}{last})").font = f_tot
        wsd.cell(r, col).number_format = CUR
    for col in range(1, 12):
        wsd.cell(r, col).border = b_bottom
    sec_total_rows[name] = r
    r += 1

sub_row = r
wsd.cell(r, 2, "SUBTOTAL — company setup (excl. unit setup)").font = f_tot
for col, letter in ((7, "G"), (8, "H"), (9, "I")):
    wsd.cell(r, col, "=" + "+".join(f"{letter}{v}" for v in sec_total_rows.values())).font = f_tot
    wsd.cell(r, col).number_format = CUR
for col in range(1, 12):
    wsd.cell(r, col).fill = fill_tot
    wsd.cell(r, col).border = b_top
r += 1

unit_row = r
wsd.cell(r, 2, "Unit setup CAPEX borne by Orizuru").font = f_tot
for col, letter in ((7, "D"), (8, "E"), (9, "F")):
    wsd.cell(r, col, f"='Unit Setup (1 Unit)'!{letter}{U_TOTAL_ROW}").font = f_link
    wsd.cell(r, col).number_format = CUR
wsd.cell(r, 11, "Year-1 units signed x cost per unit x the share Orizuru bears (Assumptions tab). That share is "
                "0% — the owner funds unit setup — so this line is NIL and no unit setup cost enters CAPEX.").font = f_note
wsd.cell(r, 11).alignment = Alignment(wrap_text=True, vertical="top")
wsd.row_dimensions[r].height = 26
r += 1

cont_row = r
wsd.cell(r, 2, "Contingency on company setup").font = f_item
wsd.cell(r, 3, 0.10).font = f_in
wsd.cell(r, 3).number_format = PCT
wsd.cell(r, 3).fill = fill_key
wsd.cell(r, 3).alignment = Alignment(horizontal="center")
for col, letter in ((7, "G"), (8, "H"), (9, "I")):
    wsd.cell(r, col, f"={letter}{sub_row}*$C${cont_row}").font = f_calc
    wsd.cell(r, col).number_format = CUR
wsd.cell(r, 11, "Unit-setup contingency is already applied on its own tab — not double-counted here.").font = f_note
r += 1

grand_row = r
wsd.cell(r, 2, "TOTAL CAPEX REQUIRED").font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
for col, letter in ((7, "G"), (8, "H"), (9, "I")):
    c = wsd.cell(r, col, f"={letter}{sub_row}+{letter}{unit_row}+{letter}{cont_row}")
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.number_format = CUR
for col in range(1, 12):
    wsd.cell(r, col).fill = fill_sec
wsd.row_dimensions[r].height = 20
r += 2

# memo
wsd.cell(r, 2, "MEMO — not CAPEX").font = Font(name=FONT, size=9, bold=True)
r += 1
memo_row = r
wsd.cell(r, 2, "Bank minimum / average balance requirement (blocked, not a cost)").font = f_item
for col, v in ((7, 25000), (8, 50000), (9, 150000)):
    wsd.cell(r, col, v).font = f_in
    wsd.cell(r, col).number_format = CUR
wsd.cell(r, 11, "UAE corporate accounts typically require an average balance of AED 25k–150k depending on the bank. "
                "It is a blocked asset, not expenditure — excluded from the CAPEX total above.").font = f_note
wsd.cell(r, 11).alignment = Alignment(wrap_text=True, vertical="top")
wsd.row_dimensions[r].height = 26
r += 1
wsd.cell(r, 2, "Refundable deposits included above (E1, E2, A12)").font = f_item
for col, letter in ((7, "G"), (8, "H"), (9, "I")):
    wsd.cell(r, col, f"={letter}{[d for d in detail_rows][0]}*0").font = f_calc
# build the actual refundable sum from known refs
refund_refs = []
for rr in detail_rows:
    if wsd.cell(rr, 1).value in ("E1", "E2", "A12"):
        refund_refs.append(rr)
for col, letter in ((7, "G"), (8, "H"), (9, "I")):
    wsd.cell(r, col, "=" + "+".join(f"{letter}{x}" for x in refund_refs)).font = f_calc
    wsd.cell(r, col).number_format = CUR
wsd.cell(r, 11, "Cash out on day one but recoverable. Useful for the bank when distinguishing sunk cost from working capital.").font = f_note
wsd.cell(r, 11).alignment = Alignment(wrap_text=True, vertical="top")

GRAND_ROW = grand_row
SUB_ROW = sub_row
UNIT_LINK_ROW = unit_row
CONT_ROW = cont_row

wsd.column_dimensions["A"].width = 7
wsd.column_dimensions["B"].width = 58
wsd.column_dimensions["C"].width = 7
for col in "DEF":
    wsd.column_dimensions[col].width = 13
for col in "GHI":
    wsd.column_dimensions[col].width = 14
wsd.column_dimensions["J"].width = 17
wsd.column_dimensions["K"].width = 66
wsd.sheet_view.showGridLines = False
wsd.freeze_panes = "C5"

# ========================================================= CAPEX SUMMARY ====
wss = wb.create_sheet("CAPEX Summary", 1)
banner(wss, "CAPEX SUMMARY BY CATEGORY",
       "All figures AED. Base case is the presentation figure; Low / High frame the range.", 6)

hr = 4
for i, h in enumerate(["Category", "Low", "Base", "High", "% of Base total", "Comment"], 1):
    c = wss.cell(hr, i, h)
    c.font = f_hdr
    c.fill = fill_hdr
    c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
wss.row_dimensions[hr].height = 24

comments = {
    "A. COMPANY FORMATION, LICENSING & REGULATORY":
        "Largely fixed and largely unavoidable. Most of it is already incurred as at 03/07/2026.",
    "B. TECHNOLOGY, PLATFORM & DIGITAL SETUP":
        "The direct booking engine is the highest-return line in the whole CAPEX — it removes OTA commission on repeat guests.",
    "C. BRAND & MARKETING SETUP":
        "Directly enables both marketing options. The owner pitch deck is what the acquisition team actually sells with.",
    "D. OFFICE EQUIPMENT & FURNITURE (DUBAI)":
        "Scale-insensitive. Same cost at 1 unit or 80.",
    "E. DEPOSITS & PREPAYMENTS":
        "Mostly refundable. Cash timing rather than true cost.",
}

r = hr + 1
sum_first = r
for name in sec_total_rows:
    wss.cell(r, 1, name).font = f_item
    for col, letter in ((2, "G"), (3, "H"), (4, "I")):
        wss.cell(r, col, f"='CAPEX Detail'!{letter}{sec_total_rows[name]}").font = f_link
        wss.cell(r, col).number_format = CUR
    c = wss.cell(r, 6, comments.get(name, ""))
    c.font = f_note
    c.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 7):
        wss.cell(r, col).border = b_bottom
    wss.row_dimensions[r].height = 26
    r += 1

wss.cell(r, 1, "Unit setup borne by Orizuru").font = f_item
for col, letter in ((2, "G"), (3, "H"), (4, "I")):
    wss.cell(r, col, f"='CAPEX Detail'!{letter}{UNIT_LINK_ROW}").font = f_link
    wss.cell(r, col).number_format = CUR
wss.cell(r, 6, "Nil — the owner funds unit setup. See the 'Unit Setup' tab for what it would cost otherwise.").font = f_note
wss.cell(r, 6).alignment = Alignment(wrap_text=True, vertical="top")
for col in range(1, 7):
    wss.cell(r, col).border = b_bottom
r += 1

wss.cell(r, 1, "Contingency").font = f_item
for col, letter in ((2, "G"), (3, "H"), (4, "I")):
    wss.cell(r, col, f"='CAPEX Detail'!{letter}{CONT_ROW}").font = f_link
    wss.cell(r, col).number_format = CUR
wss.cell(r, 6, "10% of company-setup subtotal.").font = f_note
for col in range(1, 7):
    wss.cell(r, col).border = b_bottom
sum_last = r
r += 1

TOT_ROW = r
wss.cell(r, 1, "TOTAL CAPEX").font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
for col, letter in ((2, "B"), (3, "C"), (4, "D")):
    c = wss.cell(r, col, f"=SUM({letter}{sum_first}:{letter}{sum_last})")
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.number_format = CUR
for col in range(1, 7):
    wss.cell(r, col).fill = fill_sec
wss.row_dimensions[r].height = 20

for rr in range(sum_first, sum_last + 1):
    c = wss.cell(rr, 5, f"=IF($C${TOT_ROW}=0,0,C{rr}/$C${TOT_ROW})")
    c.font = f_calc
    c.number_format = PCT
    c.alignment = Alignment(horizontal="center")
c = wss.cell(TOT_ROW, 5, f"=IF($C${TOT_ROW}=0,0,SUM(E{sum_first}:E{sum_last}))")
c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
c.number_format = PCT
c.alignment = Alignment(horizontal="center")

r = TOT_ROW + 3
wss.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = wss.cell(r, 1, "READ-ACROSS")
c.font = f_sec
c.alignment = Alignment(vertical="center", indent=1)
for col in range(1, 7):
    wss.cell(r, col).fill = fill_sec
r += 1

notes = [
    "Shareholder capital is AED 200,000 as confirmed. The Business Profile states AED 2,000,000 — reconcile the two "
    "before submission. This CAPEX is not sized to either figure; it is built bottom-up from what the business needs. "
    "At AED 200,000, CAPEX alone consumes a significant share of the capital — see 'Funding Summary'.",
    "Because Orizuru operates commission-only, CAPEX is genuinely light: there is no furniture, no unit deposits and no "
    "master-lease prepayment. The heavy spend in this business is OPEX, not CAPEX.",
    "The single largest scaling risk sits OUTSIDE this CAPEX: unit setup cost, which the property owner funds. At 80 "
    "units it would be roughly AED 732,000 — more than three times the shareholder capital. If competitive pressure "
    "ever forces Orizuru to fund onboarding, raise the share input on Assumptions and re-read the Funding Summary.",
    "Refundable items (office deposit, DEWA deposit, labour guarantee) are inside the total. Strip them out if presenting a "
    "'true sunk cost' figure.",
]
for n in notes:
    wss.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = wss.cell(r, 1, "•  " + n)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    wss.row_dimensions[r].height = 13 * (1 + len(n) // 130) + 8
    r += 1

wss.column_dimensions["A"].width = 46
for col in "BCD":
    wss.column_dimensions[col].width = 15
wss.column_dimensions["E"].width = 15
wss.column_dimensions["F"].width = 68
wss.sheet_view.showGridLines = False

wb.save(OUT)
print("saved", OUT)

"""Build Orizuru Holiday Homes LLC — OPEX workbook (AED), 12-month monthly.

Model basis (confirmed by HJ):
  - Pure management / commission-only. No unit rent, no owner utilities in OPEX.
  - Directors draw no salary.
  - Operations & service staff: 1 person @ AED 3,000/month.
  - Marketing Option 1 (fully online) and Option 2 (remote office outside UAE)
    run IN PARALLEL for the first 2 months, then one is adopted.
  - The 2% per-unit incentive is an ONGOING trailing commission on that unit's
    monthly revenue, not a one-off signing bonus.
  - AI subscriptions are a named OPEX line with their own detail tab.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/home/user/scaling-couscous/orizuru/Orizuru_OPEX.xlsx"

FONT = "Arial"
NAVY = "1F3352"
SAND = "EDE7DC"
RULE = "C7B99A"

f_title = Font(name=FONT, size=14, bold=True, color="FFFFFF")
f_sub = Font(name=FONT, size=9, color="FFFFFF")
f_sec = Font(name=FONT, size=10, bold=True, color="FFFFFF")
f_hdr = Font(name=FONT, size=9, bold=True, color="FFFFFF")
f_item = Font(name=FONT, size=10)
f_in = Font(name=FONT, size=10, color="0000FF")
f_calc = Font(name=FONT, size=10)
f_link = Font(name=FONT, size=10, color="008000")
f_tot = Font(name=FONT, size=10, bold=True)
f_note = Font(name=FONT, size=8, color="595959", italic=True)
f_white_b = Font(name=FONT, size=11, bold=True, color="FFFFFF")

fill_title = PatternFill("solid", fgColor=NAVY)
fill_sec = PatternFill("solid", fgColor="41597A")
fill_hdr = PatternFill("solid", fgColor="6B7F99")
fill_key = PatternFill("solid", fgColor="FFFF00")
fill_tot = PatternFill("solid", fgColor="DCE3EC")
fill_band = PatternFill("solid", fgColor=SAND)
fill_o1 = PatternFill("solid", fgColor="E4EFE4")
fill_o2 = PatternFill("solid", fgColor="E7EDF6")

thin = Side(style="thin", color=RULE)
b_bottom = Border(bottom=thin)
b_top = Border(top=Side(style="medium", color=NAVY))

CUR = '#,##0;(#,##0);-'
CUR2 = '#,##0.00;(#,##0.00);-'
PCT = '0.0%'
PCT2 = '0.00%'
NUM = '#,##0.0;(#,##0.0);-'

wb = Workbook()
del wb["Sheet"]


def banner(ws, title, subtitle, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(1, 1, title)
    c.font = f_title
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c = ws.cell(2, 1, subtitle)
    c.font = f_sub
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 16
    for col in range(1, width + 1):
        ws.cell(1, col).fill = fill_title
        ws.cell(2, col).fill = fill_title


def section(ws, r, text, width):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
    c = ws.cell(r, 1, text)
    c.font = f_sec
    c.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, width + 1):
        ws.cell(r, col).fill = fill_sec
    ws.row_dimensions[r].height = 18


# ================================================== AI SUBSCRIPTIONS TAB ====
wsa = wb.create_sheet("AI Subscriptions")
banner(wsa, "AI SUBSCRIPTIONS — DETAIL",
       "HJ item 5. Feeds the 'AI subscriptions' line in the OPEX model. USD converted at the pegged AED rate.", 7)

wsa.cell(4, 1, "USD/AED peg").font = f_item
c = wsa.cell(4, 2, 3.6725)
c.font = f_in
c.fill = fill_key
c.number_format = '0.0000'
wsa.cell(4, 3, "AED is pegged to USD at 3.6725. Fixed, so no FX risk on these subscriptions.").font = f_note

hr = 6
for i, h in enumerate(["Tool", "Plan", "Seats", "Price/seat (USD, monthly)",
                       "Monthly cost (AED)", "Annual cost (AED)", "What it does for Orizuru"], 1):
    c = wsa.cell(hr, i, h)
    c.font = f_hdr
    c.fill = fill_hdr
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
wsa.row_dimensions[hr].height = 30

ai = [
    ("Claude", "Max (5x)", 1, 100,
     "Core workhorse: owner proposals, management contracts, DET compliance drafting, financial reporting, guest-comms tone-of-voice."),
    ("Claude", "Pro", 2, 20,
     "Seats for the operations staff member and the marketing team lead."),
    ("ChatGPT", "Business", 2, 25,
     "Second-opinion model plus image generation for listing and social content."),
    ("Google Gemini", "AI Pro (via Workspace)", 2, 20,
     "Long-context document review and Sheets/Docs automation for owner reporting."),
    ("Perplexity", "Pro", 1, 20,
     "Live comparable-rate research: what competing units in the same tower are charging tonight."),
    ("Canva", "Pro / Teams", 2, 15,
     "Listing graphics, owner pitch decks, social content at volume without a designer."),
    ("PMS AI guest-messaging add-on", "Per-property tier", 1, 49,
     "Auto-responds to routine guest questions 24/7. Directly supports the 24/7 support promise in the Business Profile."),
    ("AI transcription / meeting notes", "Standard", 1, 18,
     "Owner calls and site visits captured without manual note-taking."),
    ("Buffer / social scheduling with AI", "Team", 1, 12,
     "Schedules the marketing team's output across channels."),
]

r = hr + 1
ai_first = r
for tool, plan, seats, usd, note in ai:
    wsa.cell(r, 1, tool).font = f_item
    wsa.cell(r, 2, plan).font = f_item
    wsa.cell(r, 3, seats).font = f_in
    wsa.cell(r, 3).alignment = Alignment(horizontal="center")
    wsa.cell(r, 4, usd).font = f_in
    wsa.cell(r, 4).number_format = CUR2
    wsa.cell(r, 5, f"=C{r}*D{r}*$B$4").font = f_calc
    wsa.cell(r, 5).number_format = CUR
    wsa.cell(r, 6, f"=E{r}*12").font = f_calc
    wsa.cell(r, 6).number_format = CUR
    n = wsa.cell(r, 7, note)
    n.font = f_note
    n.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 8):
        wsa.cell(r, col).border = b_bottom
    wsa.row_dimensions[r].height = max(24, 12 * (1 + len(note) // 72))
    r += 1
ai_last = r - 1

AI_TOTAL_ROW = r
wsa.cell(r, 1, "TOTAL — AI SUBSCRIPTIONS").font = f_white_b
wsa.cell(r, 5, f"=SUM(E{ai_first}:E{ai_last})").font = f_white_b
wsa.cell(r, 5).number_format = CUR
wsa.cell(r, 6, f"=SUM(F{ai_first}:F{ai_last})").font = f_white_b
wsa.cell(r, 6).number_format = CUR
for col in range(1, 8):
    wsa.cell(r, col).fill = fill_sec
wsa.row_dimensions[r].height = 20

r += 2
wsa.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c = wsa.cell(r, 1, "This is the Base case. In the OPEX model, the Low case assumes a trimmed stack (Claude Max + one "
                   "assistant seat only) and the High case assumes the full stack plus additional seats as the "
                   "marketing team scales. Edit seats and prices above and the Base figure flows straight through.")
c.font = f_note
c.alignment = Alignment(wrap_text=True, vertical="top")
wsa.row_dimensions[r].height = 40

wsa.column_dimensions["A"].width = 30
wsa.column_dimensions["B"].width = 22
wsa.column_dimensions["C"].width = 8
wsa.column_dimensions["D"].width = 18
wsa.column_dimensions["E"].width = 17
wsa.column_dimensions["F"].width = 17
wsa.column_dimensions["G"].width = 66
wsa.sheet_view.showGridLines = False

# ======================================================== ASSUMPTIONS TAB ===
wsx = wb.create_sheet("Assumptions", 0)
banner(wsx, "OPEX ASSUMPTIONS — ORIZURU HOLIDAY HOMES LLC",
       "All figures AED unless stated. Blue = editable input. Yellow = figure supplied by HJ or a key lever.", 7)

wsx.cell(4, 1, "SCENARIO").font = Font(name=FONT, size=11, bold=True)
c = wsx.cell(4, 2, "Base")
c.font = Font(name=FONT, size=11, bold=True, color="0000FF")
c.fill = fill_key
c.alignment = Alignment(horizontal="center")
dv = DataValidation(type="list", formula1='"Low,Base,High"', allow_blank=False)
wsx.add_data_validation(dv)
dv.add(c)
wsx.cell(4, 3, "Change this cell to Low / Base / High. Every 'Active' figure below and the whole monthly model follow it.").font = f_note

wsx.cell(5, 1, "Scenario index").font = f_note
wsx.cell(5, 2, '=MATCH($B$4,$J$4:$J$6,0)').font = f_calc
wsx.cell(5, 2).alignment = Alignment(horizontal="center")
for i, v in enumerate(["Low", "Base", "High"]):
    wsx.cell(4 + i, 10, v).font = Font(name=FONT, size=8, color="BFBFBF")

# marketing decision switch
wsx.cell(6, 1, "Option adopted from Month 3").font = Font(name=FONT, size=11, bold=True)
c = wsx.cell(6, 2, "Option 2")
c.font = Font(name=FONT, size=11, bold=True, color="0000FF")
c.fill = fill_key
c.alignment = Alignment(horizontal="center")
dv2 = DataValidation(type="list", formula1='"Option 1,Option 2,Both"', allow_blank=False)
wsx.add_data_validation(dv2)
dv2.add(c)
wsx.cell(6, 3, "Both options run in Months 1–2 per HJ. From Month 3 the model keeps whichever you select here. "
               "Select 'Both' to keep running them in parallel all year.").font = f_note

wsx.cell(7, 1, "Parallel-run months").font = f_item
c = wsx.cell(7, 2, 2)
c.font = f_in
c.fill = fill_key
c.alignment = Alignment(horizontal="center")
wsx.cell(7, 3, "Per HJ: run both marketing options together for the first 2 months to see which works.").font = f_note

SCEN_IDX = "Assumptions!$B$5"
ADOPTED = "Assumptions!$B$6"
PARALLEL = "Assumptions!$B$7"

hr = 9
for i, h in enumerate(["Assumption", "Unit", "Low", "Base", "High", "Active", "Source / note"], 1):
    c = wsx.cell(hr, i, h)
    c.font = f_hdr
    c.fill = fill_hdr
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
wsx.row_dimensions[hr].height = 24

R = {}
r = hr + 1


def sec_row(text):
    global r
    section(wsx, r, text, 7)
    r += 1


def a(key, label, unit, lo, ba, hi, note, fmt=CUR, hj=False, formula=None):
    """Add an assumption row. lo/ba/hi may be numbers or formula strings."""
    global r
    wsx.cell(r, 1, label).font = f_item
    wsx.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    wsx.cell(r, 2, unit).font = f_item
    wsx.cell(r, 2).alignment = Alignment(horizontal="center")
    for col, v in ((3, lo), (4, ba), (5, hi)):
        c = wsx.cell(r, col, v)
        is_f = isinstance(v, str) and v.startswith("=")
        c.font = f_calc if is_f else f_in
        if isinstance(v, str) and "'AI Subscriptions'" in v:
            c.font = f_link
        c.number_format = fmt
        if hj and not is_f:
            c.fill = fill_key
    c = wsx.cell(r, 6, f"=INDEX(C{r}:E{r},1,{SCEN_IDX})")
    c.font = f_tot
    c.number_format = fmt
    c.fill = fill_tot
    n = wsx.cell(r, 7, note)
    n.font = f_note
    n.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 8):
        wsx.cell(r, col).border = b_bottom
    wsx.row_dimensions[r].height = max(22, 11 * (1 + len(note) // 78))
    R[key] = r
    r += 1


sec_row("1.  REVENUE DRIVERS  —  needed to size the 2% commission, gateway fees and cost-to-revenue ratios")
a("adr", "Average Daily Rate (ADR) per unit", "AED/night", 380, 450, 550,
  "Premium 1BR short-term rental, Dubai. Drives the revenue base for the 2% marketing commission.")
a("occ", "Occupancy rate", "%", 0.65, 0.75, 0.82,
  "Dubai holiday-home market runs high; 75% is a realistic stabilised assumption.", fmt=PCT)
a("nights", "Nights per month", "nights", 30.4, 30.4, 30.4, "365 / 12.", fmt=NUM)
a("gross_unit", "Gross booking revenue per unit per month", "AED", None, None, None, "", fmt=CUR)
# fix the three formula cells for gross_unit
gr = R["gross_unit"]
for col, letter in ((3, "C"), (4, "D"), (5, "E")):
    wsx.cell(gr, col, f"={letter}{R['adr']}*{letter}{R['occ']}*{letter}{R['nights']}").font = f_calc
    wsx.cell(gr, col).number_format = CUR
wsx.cell(gr, 7, "Calculated. This is the GUEST's spend on that unit — not Orizuru's revenue.").font = f_note

a("mgmt_fee", "Management fee retained by Orizuru", "% of gross", 0.15, 0.20, 0.25,
  "Dubai holiday-home management contracts run 15–25%. This IS Orizuru's revenue under the commission-only model.",
  fmt=PCT)
a("rev_unit", "Orizuru revenue per unit per month", "AED", None, None, None, "", fmt=CUR)
rr = R["rev_unit"]
for col, letter in ((3, "C"), (4, "D"), (5, "E")):
    wsx.cell(rr, col, f"={letter}{gr}*{letter}{R['mgmt_fee']}").font = f_calc
    wsx.cell(rr, col).number_format = CUR
wsx.cell(rr, 7, "Calculated. Compare against the Business Profile's AED 2.5–4.5M turnover target — see the "
                "reconciliation note on the README tab.").font = f_note

sec_row("2.  STAFF  —  Dubai payroll")
a("dir_sal", "Directors' salary", "AED/month", 0, 0, 0,
  "Per HJ: directors will NOT draw a salary. Note for the bank — this understates the true cost of running the "
  "business, since two directors are working unpaid.", hj=True)
a("ops_hc", "Operations & service staff — headcount", "persons", 1, 1, 1,
  "Per HJ: 1 person.", fmt='0', hj=True)
a("ops_sal", "Operations & service staff — salary", "AED/month", 3000, 3000, 3000,
  "Per HJ: AED 3,000/month.", hj=True)
a("visa_amort", "Staff visa, medical insurance & EID — monthly accrual per UAE employee", "AED/month", 250, 380, 550,
  "2-year residence visa and mandatory Dubai health insurance spread monthly. Real cash cost even though it is billed "
  "in lumps.")
a("gratuity", "End-of-service gratuity accrual", "% of salary", 0.058, 0.058, 0.058,
  "UAE Labour Law: 21 days' basic pay per year of service for the first 5 years = 5.8% of salary. Legally payable — "
  "accrue it or it becomes a nasty surprise.", fmt=PCT2)
a("wps", "WPS / payroll processing fee", "AED/employee/mo", 8, 15, 25,
  "Wage Protection System is mandatory for MOHRE-sponsored staff.")

sec_row("3.  MARKETING OPTION 1  —  FULLY ONLINE  (4 hrs/day)")
a("o1_hc", "Team size", "persons", 2, 3, 4,
  "Per HJ: 2–4 persons.", fmt='0', hj=True)
a("o1_rate", "Cost per person", "AED/month", 500, 600, 700,
  "Per HJ: AED 500–700 per person.", hj=True)
a("o1_ads", "Digital ads & owner lead generation", "AED/month", 800, 1500, 2800,
  "A fully-online team without a paid lead flow will stall. This is the cost of feeding them owner leads.")
a("o1_tools", "Outreach tooling, CRM & lead database", "AED/month", 150, 280, 450,
  "LinkedIn Sales Navigator / Property Finder lead access / CRM seats.")
a("o1_mgmt", "Team coordination & QA time", "AED/month", 0, 300, 600,
  "Someone has to check the work. At 2–4 remote people this is light but not zero.")

sec_row("4.  MARKETING OPTION 2  —  REMOTE OFFICE OUTSIDE UAE  (4 hrs/day)")
a("o2_hc", "Team size", "persons", 4, 5, 6,
  "Per HJ: 4–6 persons.", fmt='0', hj=True)
a("o2_rate", "Cost per person", "AED/month", 180, 240, 300,
  "Per HJ: AED 180–300 per person.", hj=True)
a("o2_rent", "Coworking area rent", "AED/month", 1350, 1425, 1500,
  "Per HJ: AED 1,350–1,500.", hj=True)
a("o2_util", "Local internet, utilities & equipment", "AED/month", 200, 350, 550,
  "Not in HJ's original list but unavoidable — a physical room needs connectivity and machines.")
a("o2_lead", "On-site supervisor / team lead", "AED/month", 0, 500, 900,
  "Not in HJ's original list. 4–6 people in a physical room with no supervisor will drift. Recommend keeping this.")
a("o2_ads", "Digital ads & lead generation", "AED/month", 400, 900, 1600,
  "Lower than Option 1 — a larger team does more manual prospecting, so less paid volume is needed.")
a("o2_setup", "Local compliance / entity or contractor arrangement", "AED/month", 0, 400, 900,
  "Not in HJ's original list. Engaging staff outside the UAE needs a compliant contracting route. Flagging it rather "
  "than pretending it is free.")

sec_row("5.  UNIT-ACQUISITION COMMISSION  —  the 2% per HJ")
a("comm_pct", "Trailing commission on each signed unit", "% ", 0.02, 0.02, 0.02,
  "Per HJ: 2% for each unit signed, ONGOING on that unit's monthly revenue.", fmt=PCT2, hj=True)
a("comm_base", "Commission base", "—", 1, 1, 1,
  "1 = 2% of the unit's GROSS booking revenue (i.e. c.10% of Orizuru's management fee on that unit).  "
  "2 = 2% of Orizuru's management fee only (far cheaper, far weaker incentive). Change the number to switch base.",
  fmt='0', hj=True)

sec_row("6.  TECHNOLOGY, PLATFORM & AI")
a("pms_unit", "PMS / channel manager licence", "AED/unit/mo", 55, 75, 95,
  "Hostaway / Guesty style per-property pricing. Scales directly with the portfolio.")
a("pricing_unit", "Dynamic pricing tool", "AED/unit/mo", 20, 32, 45,
  "PriceLabs / Wheelhouse. Typically pays for itself many times over in RevPAR.")
a("lock_unit", "Smart-lock / access platform", "AED/unit/mo", 10, 18, 28,
  "Recurring licence on the hardware installed at CAPEX.")
a("ai_subs", "AI subscriptions", "AED/month", 800, "='AI Subscriptions'!E{}".format(AI_TOTAL_ROW), 3200,
  "Per HJ item 5. Base links live to the 'AI Subscriptions' tab. Low = trimmed stack; High = full stack plus extra "
  "seats as the marketing team scales.", hj=True)
a("workspace", "Google Workspace / business email", "AED/month", 90, 150, 260,
  "Per-seat, scales with headcount.")
a("acct_sw", "Accounting software (VAT-compliant)", "AED/month", 90, 160, 280,
  "Zoho Books / Xero. Mandatory once VAT-registered.")
a("web_host", "Website hosting & direct booking engine", "AED/month", 100, 190, 340,
  "Recurring cost on the platform built at CAPEX.")
a("voip", "VoIP, WhatsApp Business API & guest support line", "AED/month", 120, 230, 400,
  "Required to deliver the 24/7 guest support promised in the Business Profile.")

sec_row("7.  UNIT OPERATIONS  —  variable, per LIVE unit per month")
a("permit_unit", "DET holiday-home permit renewal", "AED/unit/yr", 1200, 1520, 1800,
  "Accrued monthly at 1/12. Re-chargeable to the owner in most contracts — if you re-charge, set this to 0.")
a("clean_unit", "Housekeeping & laundry — net of guest recharge", "AED/unit/mo", 0, 180, 420,
  "Cleaning fees are normally charged to the guest and pass straight through. This line is the RESIDUAL Orizuru "
  "absorbs on unsold nights, owner stays and disputes. Genuinely 0 if you recharge everything.")
a("consum_unit", "Guest consumables & amenities", "AED/unit/mo", 55, 100, 170,
  "Toiletries, coffee, water, welcome items. The Japanese-hospitality positioning makes this a brand cost, not a "
  "cost to strip out.")
a("maint_unit", "Maintenance coordination & petty repairs", "AED/unit/mo", 40, 90, 160,
  "Orizuru's coordination cost. Major repairs are billed to the owner.")
a("linen_unit", "Linen & towel replacement", "AED/unit/mo", 20, 40, 70,
  "Replacement cycle on the par stock bought at CAPEX. Real and routinely forgotten.")

sec_row("8.  OFFICE & ADMINISTRATION  —  Dubai")
a("office_rent", "Office rent / registered address", "AED/month", 1500, 2500, 4200,
  "Fronds Building – Office 02. Required to hold the trade licence.")
a("office_util", "Office DEWA, internet & chiller", "AED/month", 300, 520, 900,
  "Office only. Unit utilities sit with the owner under the commission-only model.")
a("mobile", "Mobile & telephone", "AED/month", 150, 260, 420,
  "Includes the dedicated 24/7 guest line.")
a("transport", "Transport, fuel, Salik & parking", "AED/month", 500, 1100, 2200,
  "Check-ins, inspections and owner meetings across Dubai. Rises directly with unit count — the most commonly "
  "underestimated line in this business.")
a("office_cons", "Office consumables, printing & courier", "AED/month", 100, 200, 380,
  "Contracts, DET paperwork, key courier.")
a("bank_chg", "Bank charges & account maintenance", "AED/month", 100, 175, 320,
  "Multi-currency corporate current account per the Business Profile.")
a("gateway", "Payment gateway / OTA payout fees", "% of gross collections", 0.018, 0.026, 0.032,
  "Applied to gross booking revenue passing through Orizuru's account. Excluded if the OTA pays the owner directly — "
  "set to 0 in that case.", fmt=PCT2)

sec_row("9.  PROFESSIONAL, COMPLIANCE & INSURANCE")
a("bookkeep", "Bookkeeping & management accounts", "AED/month", 500, 900, 1600,
  "Outsourced. Cheaper and more reliable than an in-house hire at this scale.")
a("vat_q", "VAT return preparation & filing", "AED/quarter", 400, 700, 1200,
  "Quarterly filing. Charged in Months 3, 6, 9 and 12.")
a("ct_pa", "Corporate Tax compliance & filing", "AED/year", 2500, 4000, 6500,
  "UAE Corporate Tax at 9% above AED 375,000 taxable profit. Compliance cost only — tax itself is not OPEX.")
a("audit_pa", "External audit", "AED/year", 3500, 6000, 10000,
  "Accrued monthly. Banks and owners increasingly require audited accounts.")
a("pro_pa", "PRO / government liaison services", "AED/year", 1500, 2800, 4500,
  "Visa renewals, licence amendments, DET filings.")
a("lic_pa", "Trade licence & DET operator permit renewal", "AED/year", 14000, 19000, 26000,
  "Accrued monthly so the renewal does not blow a hole in one month's cash. Combines DED licence and DET operator "
  "permit renewal.")
a("ins_pa", "Public liability & professional indemnity insurance", "AED/year", 3000, 5500, 9500,
  "Essential — Orizuru holds keys to and controls access to other people's property.")
a("legal_pa", "Legal, contract templates & disputes", "AED/year", 1500, 3000, 6000,
  "Management agreements, owner disputes, guest damage claims.")

sec_row("10.  CONTINGENCY")
a("conting", "Contingency", "% of total OPEX", 0.03, 0.05, 0.08,
  "First-year UAE service company. 5% is the prudent base.", fmt=PCT)

wsx.column_dimensions["A"].width = 52
wsx.column_dimensions["B"].width = 18
for col in "CDE":
    wsx.column_dimensions[col].width = 12
wsx.column_dimensions["F"].width = 14
wsx.column_dimensions["G"].width = 78
wsx.sheet_view.showGridLines = False
wsx.freeze_panes = "A10"


def A(key):
    return f"Assumptions!$F${R[key]}"


# ======================================================== OPEX MONTHLY ======
wsm = wb.create_sheet("OPEX Monthly", 1)
W = 16  # A label, B basis, C..N months, O FY total, P % of total
banner(wsm, "OPEX — MONTH BY MONTH, YEAR 1",
       "All figures AED. Commission-only model: no unit rent, no owner utilities. Driven entirely by the "
       "Assumptions tab — change the scenario there and this whole sheet moves.", W)

MC = {i: get_column_letter(2 + i) for i in range(1, 13)}   # month i -> column letter (C..N)
TOTCOL = "O"
PCTCOL = "P"

hr = 4
wsm.cell(hr, 1, "Line item").font = f_hdr
wsm.cell(hr, 2, "Basis").font = f_hdr
for i in range(1, 13):
    c = wsm.cell(hr, 2 + i, f"M{i}")
    c.font = f_hdr
    c.alignment = Alignment(horizontal="center")
wsm.cell(hr, 15, "FY1 TOTAL").font = f_hdr
wsm.cell(hr, 16, "% of OPEX").font = f_hdr
for col in range(1, W + 1):
    wsm.cell(hr, col).fill = fill_hdr
    wsm.cell(hr, col).alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
wsm.cell(hr, 1).alignment = Alignment(horizontal="left", indent=1, vertical="center")
wsm.cell(hr, 2).alignment = Alignment(horizontal="left", indent=1, vertical="center")
wsm.row_dimensions[hr].height = 22

r = hr + 1
MR = {}
_SELF = 0


def S():
    """Row number of the line currently being written — for self-referencing running totals."""
    return _SELF


def drow(key, label, basis, formula, fmt=CUR, total=True, font=None, fill=None, pct=False):
    """Write one monthly row. `formula` is a callable(i)->str or a constant."""
    global r
    wsm.cell(r, 1, label).font = font or f_item
    wsm.cell(r, 1).alignment = Alignment(indent=1, wrap_text=True, vertical="center")
    wsm.cell(r, 2, basis).font = f_note
    wsm.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="center")
    global _SELF
    _SELF = r
    for i in range(1, 13):
        v = formula(i) if callable(formula) else formula
        c = wsm.cell(r, 2 + i, v)
        c.font = font or (f_calc if isinstance(v, str) and v.startswith("=") else f_in)
        c.number_format = fmt
        if fill:
            c.fill = fill
    if total:
        c = wsm.cell(r, 15, f"=SUM(C{r}:N{r})")
        c.font = font or f_tot
        c.number_format = fmt
        c.fill = fill_tot
    for col in range(1, W + 1):
        wsm.cell(r, col).border = b_bottom
    wsm.row_dimensions[r].height = 20
    MR[key] = r
    r += 1
    return r - 1


# ---- drivers block
section(wsm, r, "DRIVERS  —  edit the yellow ramp row to change the whole model", W)
r += 1

drow("mnum", "Month number", "", lambda i: i, fmt='0', total=False, font=Font(name=FONT, size=9, bold=True))
ramp = [3, 4, 5, 6, 6, 7, 7, 8, 8, 8, 9, 9]
drow("signed", "New units signed in month", "Input — editable",
     lambda i: ramp[i - 1], fmt='0', font=f_in, fill=fill_key)
wsm.cell(MR["signed"], 15).font = f_tot
drow("cum", "Cumulative units signed", "Running total",
     lambda i: f"=C{MR['signed']}" if i == 1 else f"={MC[i-1]}{S()}+{MC[i]}{MR['signed']}",
     fmt='0', total=False)
wsm.cell(MR["cum"], 15, f"=N{MR['cum']}").font = f_tot
wsm.cell(MR["cum"], 15).number_format = '0'
wsm.cell(MR["cum"], 15).fill = fill_tot
drow("live", "Live revenue-generating units", "1-month onboarding lag",
     lambda i: 0 if i == 1 else f"={MC[i-1]}{MR['cum']}", fmt='0', total=False)
wsm.cell(MR["live"], 15, f"=N{MR['live']}").font = f_tot
wsm.cell(MR["live"], 15).number_format = '0'
wsm.cell(MR["live"], 15).fill = fill_tot

drow("gross", "Gross booking revenue (portfolio)", "Live units x gross rev/unit",
     lambda i: f"={MC[i]}{MR['live']}*{A('gross_unit')}")
drow("rev", "Orizuru management fee revenue", "Gross x management fee %",
     lambda i: f"={MC[i]}{MR['gross']}*{A('mgmt_fee')}")

drow("o1on", "Option 1 active (1/0)", "Both run M1–M2 per HJ",
     lambda i: f'=IF(OR({MC[i]}{MR["mnum"]}<={PARALLEL},{ADOPTED}="Option 1",{ADOPTED}="Both"),1,0)',
     fmt='0', total=False, font=Font(name=FONT, size=9, color="808080"))
drow("o2on", "Option 2 active (1/0)", "Both run M1–M2 per HJ",
     lambda i: f'=IF(OR({MC[i]}{MR["mnum"]}<={PARALLEL},{ADOPTED}="Option 2",{ADOPTED}="Both"),1,0)',
     fmt='0', total=False, font=Font(name=FONT, size=9, color="808080"))

r += 1

cat_first = {}
cat_last = {}
cat_order = []


def start_cat(name):
    global r
    section(wsm, r, name, W)
    r += 1
    cat_order.append(name)
    cat_first[name] = r


def end_cat(name, label):
    global r
    cat_last[name] = r - 1
    wsm.cell(r, 1, label).font = f_tot
    wsm.cell(r, 1).alignment = Alignment(indent=1)
    for i in range(1, 13):
        c = wsm.cell(r, 2 + i, f"=SUM({MC[i]}{cat_first[name]}:{MC[i]}{cat_last[name]})")
        c.font = f_tot
        c.number_format = CUR
    c = wsm.cell(r, 15, f"=SUM(C{r}:N{r})")
    c.font = f_tot
    c.number_format = CUR
    for col in range(1, W + 1):
        wsm.cell(r, col).fill = fill_tot
        wsm.cell(r, col).border = b_top
    CAT_TOTALS[name] = r
    r += 2


CAT_TOTALS = {}

# ---- A. STAFF
start_cat("A.  STAFF COSTS  —  DUBAI")
drow("dir", "Directors' salary", "Per HJ — nil",
     lambda i: f"={A('dir_sal')}")
drow("ops", "Operations & service staff", "Per HJ — 1 x AED 3,000",
     lambda i: f"={A('ops_hc')}*{A('ops_sal')}")
drow("visa", "Staff visa, medical insurance & EID (accrual)", "Per UAE employee",
     lambda i: f"={A('ops_hc')}*{A('visa_amort')}")
drow("grat", "End-of-service gratuity accrual", "5.8% of salary",
     lambda i: f"={MC[i]}{MR['ops']}*{A('gratuity')}")
drow("wps_r", "WPS / payroll processing", "Per employee",
     lambda i: f"={A('ops_hc')}*{A('wps')}")
end_cat("A.  STAFF COSTS  —  DUBAI", "TOTAL STAFF COSTS")

# ---- B. MARKETING OPTION 1
start_cat("B.  MARKETING  —  OPTION 1: FULLY ONLINE  (switches off after M2 unless adopted)")
drow("o1p", "Option 1 — team cost", "Headcount x rate",
     lambda i: f"={MC[i]}{MR['o1on']}*{A('o1_hc')}*{A('o1_rate')}", fill=fill_o1)
drow("o1a", "Option 1 — digital ads & lead generation", "",
     lambda i: f"={MC[i]}{MR['o1on']}*{A('o1_ads')}", fill=fill_o1)
drow("o1t", "Option 1 — outreach tooling & CRM", "",
     lambda i: f"={MC[i]}{MR['o1on']}*{A('o1_tools')}", fill=fill_o1)
drow("o1m", "Option 1 — coordination & QA", "",
     lambda i: f"={MC[i]}{MR['o1on']}*{A('o1_mgmt')}", fill=fill_o1)
end_cat("B.  MARKETING  —  OPTION 1: FULLY ONLINE  (switches off after M2 unless adopted)",
        "TOTAL — MARKETING OPTION 1")

# ---- C. MARKETING OPTION 2
start_cat("C.  MARKETING  —  OPTION 2: REMOTE OFFICE OUTSIDE UAE  (switches off after M2 unless adopted)")
drow("o2p", "Option 2 — team cost", "Headcount x rate",
     lambda i: f"={MC[i]}{MR['o2on']}*{A('o2_hc')}*{A('o2_rate')}", fill=fill_o2)
drow("o2r", "Option 2 — coworking area rent", "Per HJ",
     lambda i: f"={MC[i]}{MR['o2on']}*{A('o2_rent')}", fill=fill_o2)
drow("o2u", "Option 2 — local internet, utilities & equipment", "Added by Claude",
     lambda i: f"={MC[i]}{MR['o2on']}*{A('o2_util')}", fill=fill_o2)
drow("o2l", "Option 2 — on-site supervisor / team lead", "Added by Claude",
     lambda i: f"={MC[i]}{MR['o2on']}*{A('o2_lead')}", fill=fill_o2)
drow("o2a", "Option 2 — digital ads & lead generation", "",
     lambda i: f"={MC[i]}{MR['o2on']}*{A('o2_ads')}", fill=fill_o2)
drow("o2s", "Option 2 — offshore contracting / compliance", "Added by Claude",
     lambda i: f"={MC[i]}{MR['o2on']}*{A('o2_setup')}", fill=fill_o2)
end_cat("C.  MARKETING  —  OPTION 2: REMOTE OFFICE OUTSIDE UAE  (switches off after M2 unless adopted)",
        "TOTAL — MARKETING OPTION 2")

# ---- D. COMMISSION
start_cat("D.  UNIT-ACQUISITION COMMISSION  —  the 2% per HJ, ongoing")
drow("comm", "Trailing commission on signed units (2%)", "2% of live units' revenue, every month",
     lambda i: f"=IF({A('comm_base')}=1,{MC[i]}{MR['gross']},{MC[i]}{MR['rev']})*{A('comm_pct')}")
end_cat("D.  UNIT-ACQUISITION COMMISSION  —  the 2% per HJ, ongoing", "TOTAL COMMISSION")

# ---- E. TECHNOLOGY & AI
start_cat("E.  TECHNOLOGY, PLATFORM & AI")
drow("pms", "PMS / channel manager licence", "Per live unit",
     lambda i: f"={MC[i]}{MR['live']}*{A('pms_unit')}")
drow("pricing", "Dynamic pricing tool", "Per live unit",
     lambda i: f"={MC[i]}{MR['live']}*{A('pricing_unit')}")
drow("lock", "Smart-lock / access platform", "Per live unit",
     lambda i: f"={MC[i]}{MR['live']}*{A('lock_unit')}")
drow("ai", "AI subscriptions", "Per HJ — see 'AI Subscriptions' tab",
     lambda i: f"={A('ai_subs')}", fill=fill_key)
drow("gws", "Google Workspace & business email", "",
     lambda i: f"={A('workspace')}")
drow("acct", "Accounting software", "",
     lambda i: f"={A('acct_sw')}")
drow("web", "Website hosting & direct booking engine", "",
     lambda i: f"={A('web_host')}")
drow("voip_r", "VoIP, WhatsApp Business API & guest line", "24/7 support promise",
     lambda i: f"={A('voip')}")
end_cat("E.  TECHNOLOGY, PLATFORM & AI", "TOTAL TECHNOLOGY & AI")

# ---- F. UNIT OPERATIONS
start_cat("F.  UNIT OPERATIONS  —  variable with portfolio size")
drow("permit", "DET holiday-home permit renewals (accrual)", "Per live unit, 1/12 of annual",
     lambda i: f"={MC[i]}{MR['live']}*{A('permit_unit')}/12")
drow("clean", "Housekeeping & laundry — net of guest recharge", "Per live unit",
     lambda i: f"={MC[i]}{MR['live']}*{A('clean_unit')}")
drow("consum", "Guest consumables & amenities", "Per live unit",
     lambda i: f"={MC[i]}{MR['live']}*{A('consum_unit')}")
drow("maint", "Maintenance coordination & petty repairs", "Per live unit",
     lambda i: f"={MC[i]}{MR['live']}*{A('maint_unit')}")
drow("linen", "Linen & towel replacement", "Per live unit",
     lambda i: f"={MC[i]}{MR['live']}*{A('linen_unit')}")
end_cat("F.  UNIT OPERATIONS  —  variable with portfolio size", "TOTAL UNIT OPERATIONS")

# ---- G. OFFICE & ADMIN
start_cat("G.  OFFICE & ADMINISTRATION  —  DUBAI")
drow("orent", "Office rent / registered address", "",
     lambda i: f"={A('office_rent')}")
drow("outil", "Office DEWA, internet & chiller", "",
     lambda i: f"={A('office_util')}")
drow("mob", "Mobile & telephone", "",
     lambda i: f"={A('mobile')}")
drow("trans", "Transport, fuel, Salik & parking", "",
     lambda i: f"={A('transport')}")
drow("ocons", "Office consumables, printing & courier", "",
     lambda i: f"={A('office_cons')}")
drow("bank", "Bank charges & account maintenance", "",
     lambda i: f"={A('bank_chg')}")
drow("gate", "Payment gateway / OTA payout fees", "% of gross collections",
     lambda i: f"={MC[i]}{MR['gross']}*{A('gateway')}")
end_cat("G.  OFFICE & ADMINISTRATION  —  DUBAI", "TOTAL OFFICE & ADMIN")

# ---- H. PROFESSIONAL & COMPLIANCE
start_cat("H.  PROFESSIONAL, COMPLIANCE & INSURANCE")
drow("bk", "Bookkeeping & management accounts", "",
     lambda i: f"={A('bookkeep')}")
drow("vat", "VAT return preparation & filing", "Quarterly — M3, M6, M9, M12",
     lambda i: f"=IF(MOD({MC[i]}{MR['mnum']},3)=0,{A('vat_q')},0)")
drow("ct", "Corporate Tax compliance (accrual)", "1/12 of annual",
     lambda i: f"={A('ct_pa')}/12")
drow("aud", "External audit (accrual)", "1/12 of annual",
     lambda i: f"={A('audit_pa')}/12")
drow("pro", "PRO / government liaison (accrual)", "1/12 of annual",
     lambda i: f"={A('pro_pa')}/12")
drow("lic", "Trade licence & DET permit renewal (accrual)", "1/12 of annual",
     lambda i: f"={A('lic_pa')}/12")
drow("ins", "Public liability & PI insurance (accrual)", "1/12 of annual",
     lambda i: f"={A('ins_pa')}/12")
drow("leg", "Legal, contracts & disputes (accrual)", "1/12 of annual",
     lambda i: f"={A('legal_pa')}/12")
end_cat("H.  PROFESSIONAL, COMPLIANCE & INSURANCE", "TOTAL PROFESSIONAL & COMPLIANCE")

# ---- totals
section(wsm, r, "TOTAL OPERATING EXPENDITURE", W)
r += 1

sub_r = r
wsm.cell(r, 1, "OPEX before contingency").font = f_tot
wsm.cell(r, 1).alignment = Alignment(indent=1)
for i in range(1, 13):
    c = wsm.cell(r, 2 + i, "=" + "+".join(f"{MC[i]}{CAT_TOTALS[n]}" for n in cat_order))
    c.font = f_tot
    c.number_format = CUR
wsm.cell(r, 15, f"=SUM(C{r}:N{r})").font = f_tot
wsm.cell(r, 15).number_format = CUR
for col in range(1, W + 1):
    wsm.cell(r, col).border = b_bottom
r += 1

cont_r = r
wsm.cell(r, 1, "Contingency").font = f_item
wsm.cell(r, 1).alignment = Alignment(indent=1)
wsm.cell(r, 2, "% of OPEX").font = f_note
for i in range(1, 13):
    c = wsm.cell(r, 2 + i, f"={MC[i]}{sub_r}*{A('conting')}")
    c.font = f_calc
    c.number_format = CUR
wsm.cell(r, 15, f"=SUM(C{r}:N{r})").font = f_tot
wsm.cell(r, 15).number_format = CUR
for col in range(1, W + 1):
    wsm.cell(r, col).border = b_bottom
r += 1

TOT_R = r
wsm.cell(r, 1, "TOTAL OPEX").font = f_white_b
wsm.cell(r, 1).alignment = Alignment(indent=1)
for i in range(1, 13):
    c = wsm.cell(r, 2 + i, f"={MC[i]}{sub_r}+{MC[i]}{cont_r}")
    c.font = f_white_b
    c.number_format = CUR
c = wsm.cell(r, 15, f"=SUM(C{r}:N{r})")
c.font = f_white_b
c.number_format = CUR
for col in range(1, W + 1):
    wsm.cell(r, col).fill = fill_sec
wsm.row_dimensions[r].height = 20
r += 2

# ---- margin block
section(wsm, r, "SANITY CHECK  —  does the business cover its own costs?", W)
r += 1

drow("revchk", "Orizuru revenue (management fees)", "From drivers",
     lambda i: f"={MC[i]}{MR['rev']}", font=f_calc)
drow("opexchk", "Total OPEX", "",
     lambda i: f"={MC[i]}{TOT_R}", font=f_calc)
drow("ebitda", "Operating surplus / (deficit) before tax & depreciation", "Revenue less OPEX",
     lambda i: f"={MC[i]}{MR['revchk']}-{MC[i]}{MR['opexchk']}", font=f_tot)
drow("cumcash", "Cumulative operating cash position", "Running",
     lambda i: f"=C{MR['ebitda']}" if i == 1 else f"={MC[i-1]}{S()}+{MC[i]}{MR['ebitda']}",
     font=f_tot, total=False)
wsm.cell(MR["cumcash"], 15, f"=N{MR['cumcash']}").font = f_tot
wsm.cell(MR["cumcash"], 15).number_format = CUR
wsm.cell(MR["cumcash"], 15).fill = fill_tot
drow("opexpct", "OPEX as % of Orizuru revenue", "Lower is better",
     lambda i: f"=IF({MC[i]}{MR['revchk']}=0,0,{MC[i]}{MR['opexchk']}/{MC[i]}{MR['revchk']})",
     fmt=PCT, total=False, font=f_calc)
wsm.cell(MR["opexpct"], 15,
         f"=IF(O{MR['revchk']}=0,0,O{MR['opexchk']}/O{MR['revchk']})").font = f_tot
wsm.cell(MR["opexpct"], 15).number_format = PCT
wsm.cell(MR["opexpct"], 15).fill = fill_tot
drow("costunit", "OPEX per live unit", "Efficiency metric",
     lambda i: f"=IF({MC[i]}{MR['live']}=0,0,{MC[i]}{MR['opexchk']}/{MC[i]}{MR['live']})",
     total=False, font=f_calc)
wsm.cell(MR["costunit"], 15,
         f"=IF(N{MR['live']}=0,0,N{MR['opexchk']}/N{MR['live']})").font = f_tot
wsm.cell(MR["costunit"], 15).number_format = CUR
wsm.cell(MR["costunit"], 15).fill = fill_tot

# % of OPEX column
for name in cat_order:
    cr = CAT_TOTALS[name]
    c = wsm.cell(cr, 16, f"=IF($O${TOT_R}=0,0,$O{cr}/$O${TOT_R})")
    c.font = f_tot
    c.number_format = PCT
    c.alignment = Alignment(horizontal="center")

wsm.column_dimensions["A"].width = 46
wsm.column_dimensions["B"].width = 30
for i in range(1, 13):
    wsm.column_dimensions[MC[i]].width = 11
wsm.column_dimensions["O"].width = 14
wsm.column_dimensions["P"].width = 11
wsm.sheet_view.showGridLines = False
wsm.freeze_panes = "C5"

OPEX_TOT_ROW = TOT_R
MONTH_ROWS = dict(MR)
CAT_ROWS = dict(CAT_TOTALS)

# ================================================ MARKETING COMPARISON ======
wsc = wb.create_sheet("Marketing Comparison", 2)
banner(wsc, "MARKETING OPTION 1 vs OPTION 2 — THE MONTH-3 DECISION",
       "Both options run in parallel for Months 1–2 per HJ. This tab is what you use to pick one.", 7)

r = 4
section(wsc, r, "HEAD-TO-HEAD, STEADY-STATE MONTHLY COST", 7)
r += 1
for i, h in enumerate(["Component", "Option 1 — Fully online", "Option 2 — Remote office", "Difference",
                       "", "", "Comment"], 1):
    if h:
        c = wsc.cell(r, i, h)
        c.font = f_hdr
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    else:
        wsc.cell(r, i).fill = fill_hdr
wsc.row_dimensions[r].height = 26
r += 1

comp = [
    ("Team cost", f"={A('o1_hc')}*{A('o1_rate')}", f"={A('o2_hc')}*{A('o2_rate')}",
     "Option 2 buys roughly double the headcount for less money — offshore rates outside the UAE."),
    ("Coworking / premises rent", "=0", f"={A('o2_rent')}",
     "Option 1 has no premises. This is the single line that erodes Option 2's labour-cost advantage."),
    ("Local internet, utilities & equipment", "=0", f"={A('o2_util')}", "Physical office overhead."),
    ("On-site supervisor / team lead", f"={A('o1_mgmt')}", f"={A('o2_lead')}",
     "Option 1 needs light remote QA; Option 2 needs someone in the room."),
    ("Digital ads & lead generation", f"={A('o1_ads')}", f"={A('o2_ads')}",
     "Option 1 leans on paid leads; Option 2 leans on manual prospecting volume."),
    ("Outreach tooling & CRM", f"={A('o1_tools')}", "=0",
     "Option 1 is tooling-dependent by design."),
    ("Offshore contracting / compliance", "=0", f"={A('o2_setup')}",
     "Engaging staff outside the UAE needs a compliant route. Not in the original brief — flagged, not hidden."),
]
c_first = r
for label, f1, f2, note in comp:
    wsc.cell(r, 1, label).font = f_item
    wsc.cell(r, 2, f1).font = f_calc
    wsc.cell(r, 3, f2).font = f_calc
    wsc.cell(r, 4, f"=B{r}-C{r}").font = f_calc
    for col in (2, 3, 4):
        wsc.cell(r, col).number_format = CUR
    n = wsc.cell(r, 7, note)
    n.font = f_note
    n.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 8):
        wsc.cell(r, col).border = b_bottom
    wsc.row_dimensions[r].height = 24
    r += 1
c_last = r - 1

MKT_TOT = r
wsc.cell(r, 1, "TOTAL FIXED MONTHLY COST").font = f_white_b
for col, letter in ((2, "B"), (3, "C"), (4, "D")):
    c = wsc.cell(r, col, f"=SUM({letter}{c_first}:{letter}{c_last})")
    c.font = f_white_b
    c.number_format = CUR
for col in range(1, 8):
    wsc.cell(r, col).fill = fill_sec
wsc.row_dimensions[r].height = 20
r += 1

wsc.cell(r, 1, "Headcount").font = f_item
wsc.cell(r, 2, f"={A('o1_hc')}").font = f_calc
wsc.cell(r, 3, f"={A('o2_hc')}").font = f_calc
wsc.cell(r, 4, f"=B{r}-C{r}").font = f_calc
for col in (2, 3, 4):
    wsc.cell(r, col).number_format = '0'
wsc.cell(r, 7, "Both at 4 hours per day per HJ.").font = f_note
HC_ROW = r
r += 1

wsc.cell(r, 1, "Cost per head").font = f_item
wsc.cell(r, 2, f"=IF(B{HC_ROW}=0,0,B{MKT_TOT}/B{HC_ROW})").font = f_calc
wsc.cell(r, 3, f"=IF(C{HC_ROW}=0,0,C{MKT_TOT}/C{HC_ROW})").font = f_calc
wsc.cell(r, 4, f"=B{r}-C{r}").font = f_calc
for col in (2, 3, 4):
    wsc.cell(r, col).number_format = CUR
wsc.cell(r, 7, "All-in cost of one working person, including their share of premises and tooling. This is the number "
                "that actually decides it.").font = f_note
wsc.cell(r, 7).alignment = Alignment(wrap_text=True, vertical="top")
wsc.row_dimensions[r].height = 26
CPH_ROW = r
r += 2

# break-even on units
section(wsc, r, "BREAK-EVEN  —  how many units must each option sign to pay for itself?", 7)
r += 1
wsc.cell(r, 1, "Orizuru revenue per live unit per month").font = f_item
wsc.cell(r, 2, f"={A('rev_unit')}").font = f_link
wsc.cell(r, 2).number_format = CUR
wsc.cell(r, 7, "From the Assumptions tab.").font = f_note
RPU_ROW = r
r += 1
wsc.cell(r, 1, "Units needed to cover the option's fixed cost").font = f_tot
wsc.cell(r, 2, f"=IF($B${RPU_ROW}=0,0,B{MKT_TOT}/$B${RPU_ROW})").font = f_tot
wsc.cell(r, 3, f"=IF($B${RPU_ROW}=0,0,C{MKT_TOT}/$B${RPU_ROW})").font = f_tot
for col in (2, 3):
    wsc.cell(r, col).number_format = NUM
wsc.cell(r, 7, "Each option only has to sign this many units to wash its own face. Both numbers are small — which is "
                "why running both for two months is a cheap experiment, not a risk.").font = f_note
wsc.cell(r, 7).alignment = Alignment(wrap_text=True, vertical="top")
wsc.row_dimensions[r].height = 26
r += 2

section(wsc, r, "COST OF THE 2-MONTH PARALLEL RUN", 7)
r += 1
wsc.cell(r, 1, "Both options, Months 1–2 combined").font = f_item
wsc.cell(r, 2, f"=(B{MKT_TOT}+C{MKT_TOT})*{PARALLEL}").font = f_calc
wsc.cell(r, 2).number_format = CUR
wsc.cell(r, 7, "The total price of the experiment HJ described.").font = f_note
r += 1
wsc.cell(r, 1, "Incremental cost vs running the cheaper option alone").font = f_item
wsc.cell(r, 2, f"=(MAX(B{MKT_TOT},C{MKT_TOT}))*{PARALLEL}").font = f_calc
wsc.cell(r, 2).number_format = CUR
wsc.cell(r, 7, "What the second option costs you on top. Judge the experiment against this, not the gross figure.").font = f_note
wsc.cell(r, 7).alignment = Alignment(wrap_text=True, vertical="top")
r += 1
wsc.cell(r, 1, "Units the experiment must produce to justify itself").font = f_tot
wsc.cell(r, 2, f"=IF($B${RPU_ROW}=0,0,B{r-1}/$B${RPU_ROW})").font = f_tot
wsc.cell(r, 2).number_format = NUM
wsc.cell(r, 7, "If the losing option signs at least this many units in two months, the experiment paid for itself "
                "regardless of which one you keep.").font = f_note
wsc.cell(r, 7).alignment = Alignment(wrap_text=True, vertical="top")
wsc.row_dimensions[r].height = 26
r += 2

section(wsc, r, "WHAT TO MEASURE BEFORE YOU DECIDE IN MONTH 3", 7)
r += 1
decide = [
    "Units signed per option — the only metric that ultimately matters.",
    "Cost per signed unit = the option's 2-month cost / units it signed. Compare directly.",
    "Owner meetings booked per person per week — tells you whether the difference is effort or conversion.",
    "Lead-to-meeting and meeting-to-signature conversion rates, tracked separately per option.",
    "Quality of units signed: a team that signs cheap studios in weak buildings is not outperforming.",
    "Attrition and reliability — Option 2's larger offshore team carries more people risk.",
    "Language and time-zone fit with Dubai property owners. Under-weighted and it decides more deals than price does.",
]
for d in decide:
    wsc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = wsc.cell(r, 1, "•  " + d)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    wsc.row_dimensions[r].height = 13 * (1 + len(d) // 120) + 6
    r += 1
r += 1

section(wsc, r, "READ-ACROSS", 7)
r += 1
for d in [
    "Option 2 typically wins on raw cost per head even after paying coworking rent, because it buys 4–6 people for "
    "roughly what 2–4 cost online. Its risk is not cost — it is supervision and consistency.",
    "Option 1's real cost is the ads budget, not the salaries. Strip the ad spend out and it looks cheap, but a "
    "fully-online team with no lead flow signs nothing.",
    "The 2% trailing commission applies identically to both options, so it does not affect the choice — but it does "
    "mean marketing cost keeps growing with the portfolio long after the team is fixed. See the OPEX Monthly tab.",
    "Whatever you choose, the decision should be made on cost per SIGNED UNIT, not on total spend.",
]:
    wsc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = wsc.cell(r, 1, "•  " + d)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    wsc.row_dimensions[r].height = 13 * (1 + len(d) // 120) + 6
    r += 1

wsc.column_dimensions["A"].width = 44
for col in "BCD":
    wsc.column_dimensions[col].width = 17
wsc.column_dimensions["E"].width = 3
wsc.column_dimensions["F"].width = 3
wsc.column_dimensions["G"].width = 62
wsc.sheet_view.showGridLines = False

# ==================================================== OPEX SUMMARY ==========
wsu2 = wb.create_sheet("OPEX Summary", 1)
banner(wsu2, "OPEX SUMMARY — YEAR 1",
       "All figures AED. Reads live from 'OPEX Monthly'. Change the scenario on the Assumptions tab to re-cut it.", 6)

r = 4
for i, h in enumerate(["Category", "FY1 total", "Monthly average", "% of total OPEX", "", "Comment"], 1):
    if h:
        c = wsu2.cell(r, i, h)
        c.font = f_hdr
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    else:
        wsu2.cell(r, i).fill = fill_hdr
wsu2.row_dimensions[r].height = 24
r += 1

cat_comments = {
    "A.  STAFF COSTS  —  DUBAI":
        "Understated by design: directors work unpaid per HJ. One ops person for a 70–80 unit portfolio is thin — see the notes below.",
    "B.  MARKETING  —  OPTION 1: FULLY ONLINE  (switches off after M2 unless adopted)":
        "Runs M1–M2 only unless adopted from M3 on the Assumptions tab.",
    "C.  MARKETING  —  OPTION 2: REMOTE OFFICE OUTSIDE UAE  (switches off after M2 unless adopted)":
        "Runs M1–M2 only unless adopted from M3 on the Assumptions tab.",
    "D.  UNIT-ACQUISITION COMMISSION  —  the 2% per HJ, ongoing":
        "Grows every month as the portfolio grows — this is the line that scales fastest of anything here.",
    "E.  TECHNOLOGY, PLATFORM & AI":
        "Part fixed (AI, Workspace, website), part per-unit (PMS, pricing, locks). Includes HJ item 5.",
    "F.  UNIT OPERATIONS  —  variable with portfolio size":
        "Fully variable. Much of it is re-chargeable to owners or guests — set the relevant assumptions to 0 if you recharge.",
    "G.  OFFICE & ADMINISTRATION  —  DUBAI":
        "Mostly fixed. Transport is the one line that climbs with unit count.",
    "H.  PROFESSIONAL, COMPLIANCE & INSURANCE":
        "Annual costs accrued monthly so no single month takes the hit. Non-negotiable for a licensed UAE entity.",
}

s_first = r
for name in cat_order:
    label = name.split("  —  ")[0].strip()
    if "OPTION 1" in name:
        label = "B.  Marketing — Option 1 (fully online)"
    elif "OPTION 2" in name:
        label = "C.  Marketing — Option 2 (remote office)"
    elif name.startswith("D."):
        label = "D.  Unit-acquisition commission (2%)"
    wsu2.cell(r, 1, label).font = f_item
    wsu2.cell(r, 2, f"='OPEX Monthly'!O{CAT_ROWS[name]}").font = f_link
    wsu2.cell(r, 2).number_format = CUR
    wsu2.cell(r, 3, f"=B{r}/12").font = f_calc
    wsu2.cell(r, 3).number_format = CUR
    c = wsu2.cell(r, 6, cat_comments.get(name, ""))
    c.font = f_note
    c.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 7):
        wsu2.cell(r, col).border = b_bottom
    wsu2.row_dimensions[r].height = 26
    r += 1

wsu2.cell(r, 1, "Contingency").font = f_item
wsu2.cell(r, 2, f"='OPEX Monthly'!O{cont_r}").font = f_link
wsu2.cell(r, 2).number_format = CUR
wsu2.cell(r, 3, f"=B{r}/12").font = f_calc
wsu2.cell(r, 3).number_format = CUR
wsu2.cell(r, 6, "Set on the Assumptions tab.").font = f_note
for col in range(1, 7):
    wsu2.cell(r, col).border = b_bottom
s_last = r
r += 1

S_TOT = r
wsu2.cell(r, 1, "TOTAL OPEX — YEAR 1").font = f_white_b
wsu2.cell(r, 2, f"=SUM(B{s_first}:B{s_last})").font = f_white_b
wsu2.cell(r, 2).number_format = CUR
wsu2.cell(r, 3, f"=B{r}/12").font = f_white_b
wsu2.cell(r, 3).number_format = CUR
for col in range(1, 7):
    wsu2.cell(r, col).fill = fill_sec
wsu2.row_dimensions[r].height = 20
for rr in range(s_first, s_last + 1):
    c = wsu2.cell(rr, 4, f"=IF($B${S_TOT}=0,0,B{rr}/$B${S_TOT})")
    c.font = f_calc
    c.number_format = PCT
    c.alignment = Alignment(horizontal="center")
c = wsu2.cell(S_TOT, 4, f"=IF($B${S_TOT}=0,0,SUM(D{s_first}:D{s_last}))")
c.font = f_white_b
c.number_format = PCT
c.alignment = Alignment(horizontal="center")
r += 2

section(wsu2, r, "KEY METRICS — YEAR 1", 6)
r += 1
metrics = [
    ("Units signed by end of Year 1", f"='OPEX Monthly'!N{MONTH_ROWS['cum']}", '0',
     "Against the Business Profile's 70–80 unit Year-1 target."),
    ("Gross booking revenue (portfolio, guest spend)", f"='OPEX Monthly'!O{MONTH_ROWS['gross']}", CUR,
     "This is what guests pay across the portfolio — NOT Orizuru's revenue."),
    ("Orizuru revenue (management fees)", f"='OPEX Monthly'!O{MONTH_ROWS['rev']}", CUR,
     "Orizuru's actual top line under the commission-only model."),
    ("Total OPEX", f"=B{S_TOT}", CUR, ""),
    ("Operating surplus / (deficit)", f"='OPEX Monthly'!O{MONTH_ROWS['ebitda']}", CUR,
     "Before corporate tax and before any director remuneration."),
    ("Peak cash requirement (worst cumulative month)", f"=MIN('OPEX Monthly'!C{MONTH_ROWS['cumcash']}:N{MONTH_ROWS['cumcash']})", CUR,
     "The most negative the cumulative operating cash position ever gets. THIS is the working capital number to "
     "fund — not the annual OPEX figure."),
    ("OPEX as % of Orizuru revenue", f"=IF(B{r+2}=0,0,0)", PCT, ""),
    ("Exit-run-rate monthly OPEX (Month 12)", f"='OPEX Monthly'!N{OPEX_TOT_ROW}", CUR,
     "The steady-state monthly cost you carry into Year 2."),
    ("OPEX per live unit at Month 12", f"='OPEX Monthly'!N{MONTH_ROWS['costunit']}", CUR,
     "Efficiency measure. Should fall steadily as the portfolio grows."),
]
m_first = r
for label, f, fmt, note in metrics:
    wsu2.cell(r, 1, label).font = f_item
    wsu2.cell(r, 2, f).font = f_link if "'OPEX Monthly'" in f else f_calc
    wsu2.cell(r, 2).number_format = fmt
    c = wsu2.cell(r, 6, note)
    c.font = f_note
    c.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 7):
        wsu2.cell(r, col).border = b_bottom
    wsu2.row_dimensions[r].height = 24
    r += 1
# fix the ratio metric to reference the right rows
ratio_row = m_first + 6
wsu2.cell(ratio_row, 2,
          f"=IF(B{m_first+2}=0,0,B{m_first+3}/B{m_first+2})").font = f_calc
wsu2.cell(ratio_row, 2).number_format = PCT
wsu2.cell(ratio_row, 6, "Total OPEX divided by Orizuru's management fee revenue. Above 100% means Year 1 runs at a "
                        "loss — normal while the portfolio ramps.").font = f_note
wsu2.cell(ratio_row, 6).alignment = Alignment(wrap_text=True, vertical="top")
r += 1

section(wsu2, r, "WHAT I ADDED THAT WAS NOT IN YOUR BRIEF — AND WHY", 6)
r += 1
added = [
    "Staff visa, medical insurance, Emirates ID and end-of-service gratuity. AED 3,000/month is the salary, not the "
    "cost of employing someone in the UAE. The true loaded cost is roughly 45–50% higher.",
    "Trade licence and DET operator permit renewal, accrued monthly. These arrive as a single large annual bill and "
    "sink first-year cash plans that ignored them.",
    "Public liability and professional indemnity insurance. Orizuru holds keys to other people's property — this is "
    "not optional.",
    "External audit and Corporate Tax compliance. UAE Corporate Tax applies above AED 375,000 of taxable profit; the "
    "compliance cost applies regardless.",
    "Transport, fuel, Salik and parking. Check-ins and inspections across Dubai at 70–80 units. Consistently the most "
    "underestimated line in holiday-home management.",
    "Per-unit software: PMS/channel manager, dynamic pricing, smart-lock platform. These are priced per property, so "
    "they scale with the portfolio and are easy to miss when modelling a single unit.",
    "For Option 2: local utilities, an on-site supervisor, and a compliant offshore contracting route. A room with "
    "4–6 people in it costs more than rent plus wages.",
    "Contingency. First-year UAE service company with regulatory fees in motion.",
]
for d in added:
    wsu2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = wsu2.cell(r, 1, "•  " + d)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    wsu2.row_dimensions[r].height = 13 * (1 + len(d) // 118) + 6
    r += 1
r += 1

section(wsu2, r, "THREE THINGS TO LOOK AT BEFORE YOU CIRCULATE THIS", 6)
r += 1
for d in [
    "ONE OPERATIONS PERSON FOR 70–80 UNITS IS NOT ENOUGH. At AED 3,000/month you have one person handling check-ins, "
    "guest support, housekeeping coordination, maintenance and owner reporting across a portfolio spread over Dubai. "
    "Either headcount rises through the year, or check-in and housekeeping get outsourced per unit. The model runs "
    "with your figure as instructed — but budget for a second and third hire around Months 4 and 8.",
    "REVENUE RECOGNITION vs THE BUSINESS PROFILE. The profile projects AED 2.5–4.5M Year-1 turnover. Under a pure "
    "commission-only model at 70–80 units, Orizuru's own revenue is materially lower — the AED 2.5–4.5M is closer to "
    "gross booking value passing through. Agree with your accountant and the bank which figure you are presenting "
    "before submission; they are very different numbers.",
    "THE 2% IS ONGOING, NOT ONE-OFF. As instructed, it applies to every live unit every month, forever. By Month 12 "
    "it is one of the larger single lines in the model and it never stops growing. If the intent was a one-off "
    "signing incentive, tell me and I will re-cut it — the difference over three years is very large.",
]:
    wsu2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = wsu2.cell(r, 1, "•  " + d)
    c.font = Font(name=FONT, size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    wsu2.row_dimensions[r].height = 13 * (1 + len(d) // 118) + 8
    r += 1

wsu2.column_dimensions["A"].width = 46
wsu2.column_dimensions["B"].width = 17
wsu2.column_dimensions["C"].width = 17
wsu2.column_dimensions["D"].width = 15
wsu2.column_dimensions["E"].width = 3
wsu2.column_dimensions["F"].width = 60
wsu2.sheet_view.showGridLines = False

# ============================================================ README ========
wsr = wb.create_sheet("README", 0)
banner(wsr, "ORIZURU HOLIDAY HOMES LLC — OPERATING EXPENDITURE (OPEX)",
       "Dubai, UAE  |  Trade Licence 1634278  |  12-month monthly model  |  All figures in AED", 6)

r = 4
blocks = [
    ("SEC", "HOW TO USE THIS WORKBOOK"),
    ("T", "Tab", "What it does"),
    ("R", "Assumptions", "Every input, at Low / Base / High. Set the SCENARIO cell (B4) and the whole model re-cuts. "
                          "Set 'Option adopted from Month 3' (B6) to switch marketing options."),
    ("R", "OPEX Summary", "One-page Year-1 view: category totals, key metrics, and the notes you should read first."),
    ("R", "OPEX Monthly", "The model. 12 months across, every cost line down. Edit the yellow unit-ramp row to change "
                          "the growth path."),
    ("R", "Marketing Comparison", "Option 1 vs Option 2 head-to-head, plus what to measure before deciding in Month 3."),
    ("R", "AI Subscriptions", "HJ item 5 broken out tool by tool. Feeds the AI line in the model."),
    ("", ""),
    ("SEC", "THE FOUR LEVERS THAT MOVE EVERYTHING"),
    ("B", "1.", "Assumptions!B4 — SCENARIO (Low / Base / High)."),
    ("B", "2.", "Assumptions!B6 — which marketing option is adopted from Month 3 (or 'Both' to keep running in parallel)."),
    ("B", "3.", "OPEX Monthly, yellow row — new units signed each month. Currently ramps to 80 units by Month 12, "
                "matching the Business Profile's 70–80 target. Set every month to 0 except M1=1 for a single-unit view."),
    ("B", "4.", "Assumptions!F column — the active value of any individual line. Override any single input directly in "
                "its Low/Base/High cells."),
    ("", ""),
    ("SEC", "MODEL BASIS — CONFIRMED WITH HJ"),
    ("B", "•", "PURE MANAGEMENT / COMMISSION-ONLY. Orizuru does not master-lease. There is NO unit rent, no unit DEWA "
                "and no unit internet anywhere in this OPEX — those costs sit with the property owner."),
    ("B", "•", "Directors draw NO salary."),
    ("B", "•", "Operations & service staff: 1 person at AED 3,000/month."),
    ("B", "•", "Both marketing options run in parallel for Months 1–2, then one is adopted."),
    ("B", "•", "The 2% per signed unit is an ONGOING trailing commission on that unit's monthly revenue, applied every "
                "month for as long as the unit is live."),
    ("B", "•", "AI subscriptions are broken out on their own tab."),
    ("B", "•", "Domain (AED 387, 3 years) and company stamp (AED 137) are ONE-OFF costs and sit in the CAPEX workbook, "
                "not here."),
    ("", ""),
    ("SEC", "COLOUR LEGEND"),
    ("L", "Blue text", "Hardcoded input — edit freely."),
    ("L", "Yellow fill", "Key lever, or a figure supplied by HJ."),
    ("L", "Black text", "Formula. Do not overwrite."),
    ("L", "Green text", "Link to another sheet in this workbook."),
    ("", ""),
    ("SEC", "WHAT IS DELIBERATELY NOT IN HERE"),
    ("B", "•", "OTA commission (Airbnb / Booking.com, 15–18%). Under a commission-only model this is deducted from "
                "gross booking revenue before the owner's split — it is not an Orizuru operating cost. Add it only if "
                "your management contracts make Orizuru bear it."),
    ("B", "•", "Tourism Dirham. Collected from the guest and remitted to DET. Pure pass-through."),
    ("B", "•", "Corporate Tax itself (9% above AED 375,000 taxable profit). It is a tax on profit, not OPEX. The "
                "compliance/filing cost IS included."),
    ("B", "•", "Depreciation on the CAPEX. Non-cash — belongs in the P&L, not in an OPEX cash budget."),
    ("B", "•", "Unit furnishing and setup. Owner-funded under this model; the operating layer Orizuru does fund is in "
                "the CAPEX workbook."),
]
for row in blocks:
    if row[0] == "":
        r += 1
        continue
    if row[0] == "SEC":
        section(wsr, r, row[1], 6)
    elif row[0] == "T":
        wsr.cell(r, 1, row[1]).font = f_hdr
        wsr.cell(r, 2, row[2]).font = f_hdr
        for col in range(1, 7):
            wsr.cell(r, col).fill = fill_hdr
    else:
        wsr.cell(r, 1, row[1]).font = Font(name=FONT, size=10, bold=True)
        wsr.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = wsr.cell(r, 2, row[2])
        c.font = f_item
        c.alignment = Alignment(wrap_text=True, vertical="top")
        wsr.row_dimensions[r].height = max(14, 13 * (1 + len(row[2]) // 92))
        if row[0] == "L":
            if "Blue" in row[1]:
                wsr.cell(r, 1).font = Font(name=FONT, size=10, bold=True, color="0000FF")
            if "Green" in row[1]:
                wsr.cell(r, 1).font = Font(name=FONT, size=10, bold=True, color="008000")
            if "Yellow" in row[1]:
                wsr.cell(r, 1).fill = fill_key
    r += 1

wsr.column_dimensions["A"].width = 22
for col in "BCDEF":
    wsr.column_dimensions[col].width = 18
wsr.sheet_view.showGridLines = False

wb.move_sheet("README", offset=-10)
wb.save(OUT)
print("saved", OUT)
print("sheets:", wb.sheetnames)

"""Independent recomputation of the model, straight from the workbook's cells.

LibreOffice cannot run in this container, so instead of trusting a recalc we
re-derive every headline number in plain Python from the SAME input cells the
workbook uses, and print them. If these look right, the workbook's formulas —
which were verified structurally by check_model.py — will produce them too.
"""

from openpyxl import load_workbook

PATH = "Orizuru_CAPEX_OPEX_Model.xlsx"
wb = load_workbook(PATH)

A = wb["Assumptions"]
M = wb["OPEX Monthly"]
D = wb["CAPEX Detail"]
U = wb["Unit Setup (1 Unit)"]

SCEN_COL = {"Low": 3, "Base": 4, "High": 5}


def assumptions(scenario):
    """label -> value, for the given scenario column."""
    col = SCEN_COL[scenario]
    out = {}
    for r in range(10, A.max_row + 1):
        label = A.cell(r, 1).value
        if not label:
            continue
        v = A.cell(r, col).value
        if isinstance(v, (int, float)):
            out[label] = v
        elif isinstance(v, str) and v.startswith("='AI Subscriptions'"):
            out[label] = ai_total()   # Base case links live to the AI tab
    return out


def ai_total():
    """Recompute the AI subscription stack the way the AI tab's formulas do."""
    S = wb["AI Subscriptions"]
    peg = S["B4"].value
    t = 0.0
    for r in range(7, S.max_row + 1):
        seats, usd = S.cell(r, 3).value, S.cell(r, 4).value
        if isinstance(seats, (int, float)) and isinstance(usd, (int, float)):
            t += seats * usd * peg
    return t


def ramp():
    for r in range(5, 25):
        if M.cell(r, 1).value == "New units signed in month":
            return [M.cell(r, 2 + i).value for i in range(1, 13)]
    raise SystemExit("ramp row not found")


def run(scenario, adopted="Option 2", parallel=2):
    a = assumptions(scenario)
    gross_unit = a["Average Daily Rate (ADR) per unit"] * a["Occupancy rate"] * a["Nights per month"]
    rev_unit = gross_unit * a["Management fee retained by Orizuru"]

    signed = ramp()
    cum, live = [], []
    c = 0
    for i in range(12):
        c += signed[i]
        cum.append(c)
        live.append(0 if i == 0 else cum[i - 1])

    rows = {}
    tot = [0.0] * 12

    def add(name, fn):
        vals = [fn(i) for i in range(12)]
        rows[name] = vals
        for i in range(12):
            tot[i] += vals[i]

    gross = [live[i] * gross_unit for i in range(12)]
    rev = [g * a["Management fee retained by Orizuru"] for g in gross]

    o1 = [1 if (i + 1) <= parallel or adopted in ("Option 1", "Both") else 0 for i in range(12)]
    o2 = [1 if (i + 1) <= parallel or adopted in ("Option 2", "Both") else 0 for i in range(12)]

    # A. staff
    ops = a["Operations & service staff — headcount"] * a["Operations & service staff — salary"]
    add("A. Staff", lambda i: (
        a["Directors' salary"]
        + ops
        + a["Operations & service staff — headcount"] * a["Staff visa, medical insurance & EID — monthly accrual per UAE employee"]
        + ops * a["End-of-service gratuity accrual"]
        + a["Operations & service staff — headcount"] * a["WPS / payroll processing fee"]))

    # B / C. marketing — the two option blocks share label text, so read by row position
    def by_row(label, occurrence):
        hits = [r for r in range(10, A.max_row + 1) if A.cell(r, 1).value == label]
        return A.cell(hits[occurrence], SCEN_COL[scenario]).value

    o1_hc, o1_rate = by_row("Team size", 0), by_row("Cost per person", 0)
    o2_hc, o2_rate = by_row("Team size", 1), by_row("Cost per person", 1)

    rows["B. Marketing Opt 1"] = [o1[i] * (
        o1_hc * o1_rate
        + a["Digital ads & owner lead generation"]
        + a["Outreach tooling, CRM & lead database"]
        + a["Team coordination & QA time"]) for i in range(12)]

    rows["C. Marketing Opt 2"] = [o2[i] * (
        o2_hc * o2_rate
        + a["Coworking area rent"]
        + a["Local internet, utilities & equipment"]
        + a["On-site supervisor / team lead"]
        + a["Digital ads & lead generation"]
        + a["Local compliance / entity or contractor arrangement"]) for i in range(12)]

    rows["D. Commission (2%)"] = [gross[i] * a["Trailing commission on each signed unit"] for i in range(12)]

    rows["E. Technology & AI"] = [
        live[i] * (a["PMS / channel manager licence"] + a["Dynamic pricing tool"] + a["Smart-lock / access platform"])
        + a["AI subscriptions"] + a["Google Workspace / business email"] + a["Accounting software (VAT-compliant)"]
        + a["Website hosting & direct booking engine"] + a["VoIP, WhatsApp Business API & guest support line"]
        for i in range(12)]

    rows["F. Unit operations"] = [
        live[i] * (a["DET holiday-home permit renewal"] / 12
                   + a["Housekeeping & laundry — net of guest recharge"]
                   + a["Guest consumables & amenities"]
                   + a["Maintenance coordination & petty repairs"]
                   + a["Linen & towel replacement"])
        for i in range(12)]

    rows["G. Office & admin"] = [
        a["Office rent / registered address"] + a["Office DEWA, internet & chiller"] + a["Mobile & telephone"]
        + a["Transport, fuel, Salik & parking"] + a["Office consumables, printing & courier"]
        + a["Bank charges & account maintenance"] + gross[i] * a["Payment gateway / OTA payout fees"]
        for i in range(12)]

    rows["H. Professional & compliance"] = [
        a["Bookkeeping & management accounts"]
        + (a["VAT return preparation & filing"] if (i + 1) % 3 == 0 else 0)
        + a["Corporate Tax compliance & filing"] / 12 + a["External audit"] / 12
        + a["PRO / government liaison services"] / 12
        + a["Trade licence & DET operator permit renewal"] / 12
        + a["Public liability & professional indemnity insurance"] / 12
        + a["Legal, contract templates & disputes"] / 12
        for i in range(12)]

    sub = [sum(rows[k][i] for k in rows) for i in range(12)]
    cont = [s * a["Contingency"] for s in sub]
    total = [sub[i] + cont[i] for i in range(12)]

    surplus = [rev[i] - total[i] for i in range(12)]
    cumcash, running = [], 0.0
    for s in surplus:
        running += s
        cumcash.append(running)

    return dict(a=a, gross_unit=gross_unit, rev_unit=rev_unit, cum=cum, live=live,
                gross=gross, rev=rev, rows=rows, sub=sub, cont=cont, total=total,
                surplus=surplus, cumcash=cumcash)


def capex_totals():
    """Sum the CAPEX Detail grid and the unit sheet the same way the workbook does."""
    out = {}
    for name, col in (("Low", 4), ("Base", 5), ("High", 6)):
        s = 0.0
        for r in range(5, D.max_row + 1):
            ref = D.cell(r, 1).value
            qty = D.cell(r, 3).value
            unit = D.cell(r, col).value
            if isinstance(ref, str) and len(ref) <= 4 and isinstance(qty, (int, float)) \
                    and isinstance(unit, (int, float)):
                s += qty * unit
        out[name] = s
    # unit setup
    for name, col in (("Low", 4), ("Base", 5), ("High", 6)):
        s = 0.0
        for r in range(5, 40):
            ref = U.cell(r, 1).value
            qty = U.cell(r, 3).value
            unit = U.cell(r, col).value
            if isinstance(ref, str) and ref.startswith("U") and isinstance(qty, (int, float)) \
                    and isinstance(unit, (int, float)):
                s += qty * unit
        out["unit_" + name] = s * 1.10  # 10% contingency applied on that tab
    return out


def money(x):
    return f"{x:,.0f}"


print("=" * 84)
print("CAPEX")
print("=" * 84)
cx = capex_totals()
for s in ("Low", "Base", "High"):
    company = cx[s]
    unit = cx["unit_" + s]
    cont = company * 0.10
    print(f"  {s:<5}  company setup {money(company):>10}   +10% conting {money(cont):>9}"
          f"   + 1 unit {money(unit):>8}   =  TOTAL {money(company + cont + unit):>10}")

print()
print("=" * 84)
print("OPEX  —  Year 1, adopted = Option 2 from Month 3")
print("=" * 84)
for scen in ("Low", "Base", "High"):
    m = run(scen)
    print(f"\n--- {scen.upper()} ---")
    print(f"  Gross booking rev / unit / month : {money(m['gross_unit'])}")
    print(f"  Orizuru revenue / unit / month   : {money(m['rev_unit'])}")
    print(f"  Units signed by M12              : {m['cum'][-1]}")
    print(f"  Gross booking revenue FY1        : {money(sum(m['gross']))}")
    print(f"  Orizuru revenue FY1              : {money(sum(m['rev']))}")
    for k in m["rows"]:
        print(f"    {k:<32} {money(sum(m['rows'][k])):>12}")
    print(f"    {'Contingency':<32} {money(sum(m['cont'])):>12}")
    print(f"  TOTAL OPEX FY1                   : {money(sum(m['total']))}")
    print(f"  Monthly avg / M12 run-rate       : {money(sum(m['total'])/12)} / {money(m['total'][-1])}")
    print(f"  Operating surplus/(deficit) FY1  : {money(sum(m['surplus']))}")
    print(f"  Peak cumulative cash deficit     : {money(min(0, min(m['cumcash'])))}")
    if scen == "Base":
        base = m

print()
print("=" * 84)
print("BASE  —  month by month")
print("=" * 84)
hdr = "  " + "".join(f"{'M'+str(i+1):>9}" for i in range(12))
print(f"{'':<26}{hdr}")
for label, series in (("Live units", base["live"]), ("Orizuru revenue", base["rev"]),
                      ("Total OPEX", base["total"]), ("Surplus/(deficit)", base["surplus"]),
                      ("Cumulative cash", base["cumcash"])):
    print(f"{label:<26}  " + "".join(f"{money(v):>9}" for v in series))

print()
print("=" * 84)
print("MARKETING OPTIONS  —  steady-state monthly cost, Base")
print("=" * 84)
b = run("Base", adopted="Both")
print(f"  Option 1 (fully online)  : {money(b['rows']['B. Marketing Opt 1'][-1])} / month")
print(f"  Option 2 (remote office) : {money(b['rows']['C. Marketing Opt 2'][-1])} / month")
print(f"  2-month parallel run     : {money((b['rows']['B. Marketing Opt 1'][0] + b['rows']['C. Marketing Opt 2'][0]) * 2)}")
print(f"  Units to break even on the parallel run: "
      f"{max(b['rows']['B. Marketing Opt 1'][0], b['rows']['C. Marketing Opt 2'][0]) * 2 / base['rev_unit']:.1f}")

print()
print("=" * 84)
print("FUNDING REQUIREMENT  —  Base, 20% buffer")
print("=" * 84)
capex_total = cx["Base"] * 1.10 + cx["unit_Base"]
peak = abs(min(0, min(base["cumcash"])))
buf = (peak + capex_total) * 0.20
print(f"  CAPEX                    : {money(capex_total)}")
print(f"  Peak operating cash need : {money(peak)}")
print(f"  Buffer (20%)             : {money(buf)}")
print(f"  TOTAL FUNDING REQUIRED   : {money(capex_total + peak + buf)}")
print(f"  Shareholder capital      : {money(200000)}")
print(f"  Headroom/(shortfall)     : {money(200000 - (capex_total + peak + buf))}")
